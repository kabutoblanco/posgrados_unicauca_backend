from django.shortcuts import render
from rest_framework import viewsets, generics
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import APIException
from rest_framework.status import (
    HTTP_400_BAD_REQUEST,
    HTTP_404_NOT_FOUND,
    HTTP_200_OK,
    HTTP_201_CREATED,
    HTTP_202_ACCEPTED
)

from .backends import *
from .models import *
from .serializers import *
from .email import *

from a_students_app.models import Program
from a_students_app.models import Student, Enrrollment
from c_tracking_app.models import TestDirector, TestCoordinator
from d_information_management_app.models import Institution, InvestigationLine, Professor, City, Country, ManageInvestLine, WorksDepartm
from d_accounts_app.models import User

import json

from django.http import HttpResponse
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from django.core.serializers.json import DjangoJSONEncoder
from django.db.models import Avg, Sum, Count

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus.paragraph import Paragraph
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle

from django.db.models.signals import post_save, pre_save
from .signal import *

post_save.connect(save_or_send_activity, sender=Lecture)
post_save.connect(save_or_send_activity, sender=Publication)
post_save.connect(save_or_send_activity, sender=ProjectCourse)
post_save.connect(save_or_send_activity, sender=ResearchStays)
post_save.connect(save_or_send_activity, sender=PresentationResults)
post_save.connect(save_or_send_activity, sender=ParticipationProjects)

pre_save.connect(save_or_send_activity1, sender=Lecture)
pre_save.connect(save_or_send_activity1, sender=Publication)
pre_save.connect(save_or_send_activity1, sender=ProjectCourse)
pre_save.connect(save_or_send_activity1, sender=ResearchStays)
pre_save.connect(save_or_send_activity1, sender=PresentationResults)
pre_save.connect(save_or_send_activity1, sender=ParticipationProjects)

class ReportTest(APIView):

    def get(self, request, *args, **kwargs):

        def celdas(ws, controller):
            for i in range(2, 9):
                ws.cell(row=controller,column=i).alignment = Alignment(horizontal="center")
                ws.cell(row=controller,column=i).border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                                top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
                ws.cell(row =controller, column=i).font = Font(name='Arial',size=8)
        
        queryProfessors = Professor.objects.filter(is_internal=True)
        queryInitialPeriod = Enrrollment.objects.all().values('period').order_by('period')[:1]
        queryFinalPeriod = Enrrollment.objects.all().values('period').order_by('-period')[:1]
        queryDirectors = WorksDepartm.objects.filter(laboral_category="planta", professor__is_internal=True) | WorksDepartm.objects.filter(laboral_category="catedra", professor__is_internal=True)
        
        InitialPeriod = 0
        FinalPeriod = 0
        if (len(queryInitialPeriod) > 0  and len(queryFinalPeriod) > 0):
            InitialPeriod = queryInitialPeriod[0]["period"].split("-")[0]
            FinalPeriod = queryFinalPeriod[0]["period"].split("-")[0]
        queryStudentsProfessor = StudentProfessor.objects.all()
    

        if(kwargs["type"]==1):#Format xlsx request.data["tipo"]==1
            wb = Workbook()
            controller = 6
            ws = wb.active
            ws.title = '4.a.'

            #region Encabezados
            #Create title in the sheet
            ws['A1'].alignment = Alignment(horizontal="center",vertical="center")
            ws['A1'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
            top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['A1'].font = Font(name = 'Arial',size =8, bold =True)
            ws['A1'] = "CARACTERÍSTICA"

            ws['B1'].alignment = Alignment(horizontal="left",vertical="center")
            ws['B1'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
            top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['B1'].font = Font(name = 'Arial',size =8, bold =True)
            ws['B1'] = "7. Relación estudiante / tutor."

            ws['A2'].alignment = Alignment(horizontal="right",vertical="center")
            ws['A2'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
            top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['A2'].font = Font(name = 'Arial',size =8)
            ws['A2'] = "Indicador"

            ws['B2'].alignment = Alignment(horizontal="left",vertical="center")
            ws['B2'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
            top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['B2'].font = Font(name = 'Arial',size =8)
            ws['B2'] = "a. Número de estudiantes por tutor (sólo profesores de TC y habilitados para dirigir tesis)."

            ws['B3'].alignment = Alignment(horizontal="left",vertical="center")
            ws['B3'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
            top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['B3'].font = Font(name = 'Arial',size =8)
            ws['B3'] = "c. Número de tutores / asesores externos."

            #Change the characteristics of cells
            ws.merge_cells('B1:F1')
            ws.merge_cells('B2:F2')
            ws.merge_cells('B3:F3')

            #ws.row_dimensions[5].height = 25
            ws.column_dimensions['A'].width = 16
            ws.column_dimensions['B'].width = 14
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 17
            ws.column_dimensions['E'].width = 25
            ws.column_dimensions['F'].width = 30
            ws.column_dimensions['G'].width = 25
            ws.column_dimensions['H'].width = 17
            ws.column_dimensions['I'].width = 4
            ws.column_dimensions['J'].width = 28
            ws.column_dimensions['K'].width = 6
            ws.column_dimensions['K'].width = 6

            #Create header

            ws['B5'].alignment = Alignment(horizontal="center",vertical="center")
            ws['B5'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
            top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['B5'].fill = PatternFill(start_color = 'D7D7D7',end_color ='D7D7D7', fill_type='solid')
            ws['B5'].font = Font(name = 'Arial',size =8, bold =True)
            ws['B5']='Profesor'

            ws['C5'].alignment = Alignment(horizontal="center",vertical="center")
            ws['C5'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['C5'].fill = PatternFill(start_color = 'D7D7D7',end_color ='D7D7D7', fill_type='solid')
            ws['C5'].font = Font(name = 'Arial',size =8, bold =True)
            ws['C5']='Habilitado para dirigir'
                    
            ws['D5'].alignment = Alignment(horizontal="center",vertical="center")
            ws['D5'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['D5'].fill = PatternFill(start_color = 'D7D7D7',end_color ='D7D7D7', fill_type='solid')
            ws['D5'].font = Font(name = 'Arial',size =8, bold =True)
            ws['D5']='Número de Tutorías'

            ws['E5'].alignment = Alignment(horizontal="center",vertical="center")
            ws['E5'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['E5'].fill = PatternFill(start_color = 'D7D7D7',end_color ='D7D7D7', fill_type='solid')
            ws['E5'].font = Font(name = 'Arial',size =8, bold =True)
            ws['E5']='Estudiante'

            ws['F5'].alignment = Alignment(horizontal="center",vertical="center")
            ws['F5'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['F5'].fill = PatternFill(start_color = 'D7D7D7',end_color ='D7D7D7', fill_type='solid')
            ws['F5'].font = Font(name = 'Arial',size =8, bold =True)
            ws['F5']='Asesor Externo'
            
            ws['G5'].alignment = Alignment(horizontal="center",vertical="center")
            ws['G5'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['G5'].fill = PatternFill(start_color = 'D7D7D7',end_color ='D7D7D7', fill_type='solid')
            ws['G5'].font = Font(name = 'Arial',size =8, bold =True)
            ws['G5']='Codirector'

            ws['H5'].alignment = Alignment(horizontal="center",vertical="center")
            ws['H5'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['H5'].fill = PatternFill(start_color = 'D7D7D7',end_color ='D7D7D7', fill_type='solid')
            ws['H5'].font = Font(name = 'Arial',size =8, bold =True)
            ws['H5']='Modalidad Dirección'

            ws['J5'].alignment = Alignment(horizontal="left",vertical="center")
            ws['J5'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['J5'].font = Font(name = 'Arial',size =8, bold =True)
            ws['J5']='Número de estudiantes por tutor:'

            ws['K5'].alignment = Alignment(horizontal="right",vertical="center")
            ws['K5'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['K5'].fill = PatternFill(start_color = '78F293',end_color ='78F293', fill_type='solid')
            ws['K5'].font = Font(name = 'Arial',size =8, bold =True)

            ws['J6'].alignment = Alignment(horizontal="left",vertical="center")
            ws['J6'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['J6'].font = Font(name = 'Arial',size =8, bold =True)
            ws['J6']='Trabajos de Grado:'

            ws['K6'].alignment = Alignment(horizontal="right",vertical="center")
            ws['K6'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['K6'].fill = PatternFill(start_color = '78F293',end_color ='78F293', fill_type='solid')
            ws['K6'].font = Font(name = 'Arial',size =8, bold =True, color='FFFF0000')

            ws['J4'].alignment = Alignment(horizontal="left",vertical="center")
            ws['J4'].font = Font(name = 'Arial',size =8, bold =True, color='FFFF0000')

            ws['J7'].alignment = Alignment(horizontal="left",vertical="center")
            ws['J7'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['J7'].font = Font(name = 'Arial',size =8, bold =True)
            ws['J7']='Con tutores/asesores externos:'

            ws['K7'].alignment = Alignment(horizontal="right",vertical="center")
            ws['K7'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['K7'].fill = PatternFill(start_color = '78F293',end_color ='78F293', fill_type='solid')
            ws['K7'].font = Font(name = 'Arial',size =8, bold =True, color='FFFF0000')

            ws['L7'].alignment = Alignment(horizontal="right",vertical="center")
            ws['L7'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['L7'].fill = PatternFill(start_color = '78F293',end_color ='78F293', fill_type='solid')
            ws['L7'].font = Font(name = 'Arial',size =8, bold =True)
    #endregion
            countHabDirigir = 0
            countTutorias = 0
            countAsesoresExternos = 0
            countPrincipales = 0
            for q in queryProfessors:
                countEstProf = 0    
                defController = controller
                
                if(queryStudentsProfessor.count() == 0 ):
                    celdas(ws, controller)

                for q2 in queryStudentsProfessor:
                    celdas(ws, controller)
                    if (q.user.id == q2.professor.user.id ):
                        ws.cell(row =controller, column=5).value = q2.student.user.first_name +" "+ q2.student.user.last_name
                        countEstProf +=1

                        varRol = ''
                        nameAsesorExterno = ''
                        varModalidad = 'Co-director'
                        if(q2.rol == 1):
                            varModalidad = 'Principal'
                            countPrincipales +=1
                            for q3 in queryStudentsProfessor:
                                if(q2.student.user.id == q3.student.user.id and q3.rol ==2):
                                    if(q3.professor.is_internal == True):
                                        varRol = q3.professor.user.first_name +" "+ q3.professor.user.last_name
                                    else:
                                        nameAsesorExterno = q3.professor.user.first_name +" "+ q3.professor.user.last_name + "(extranjero)"
                                        countAsesoresExternos+=1
                        
                        ws.cell(row =controller, column=6).value = nameAsesorExterno
                        ws.cell(row =controller, column=7).value = varRol
                        ws.cell(row =controller, column=8).value = varModalidad

                        controller +=1

                if(countEstProf == 0):
                    controller +=1
                
                ws.merge_cells('B'+str(defController) +':B'+str(controller-1))
                ws.merge_cells('C'+str(defController) +':C'+str(controller-1))
                ws.merge_cells('D'+str(defController) +':D'+str(controller-1))

                ws.cell(row =defController, column=2).value = q.user.first_name +" "+ q.user.last_name

                isDirector = 0
                for q2 in queryDirectors:
                    if(q.id == q2.professor.id):
                        isDirector = 1
                        break
                        
                ws.cell(row =defController, column=3).value = str(isDirector)
                countHabDirigir += isDirector

                ws.cell(row =defController, column=4).value = str(countEstProf)
                countTutorias += countEstProf

            ws.cell(row=controller,column=2).border = Border(top=Side(border_style="thin"))

            ws['J4']="Período: "+str(InitialPeriod)+" - "+str(FinalPeriod)
            varEstXTutor = 0

            ws.cell(row=controller,column=3).alignment = Alignment(horizontal="center")
            ws.cell(row=controller,column=3).border = Border(top=Side(border_style="thin"))
            ws.cell(row =controller, column=3).font = Font(name='Arial',size=8, bold =True)
            ws.cell(row =controller, column=3).value = str(countHabDirigir)
        
            ws.cell(row=controller,column=4).alignment = Alignment(horizontal="center")
            ws.cell(row=controller,column=4).border = Border(top=Side(border_style="thin"))
            ws.cell(row =controller, column=4).font = Font(name='Arial',size=8, bold =True)
            ws.cell(row =controller, column=4).value = str(countTutorias)

            if(countHabDirigir > 0):
                varEstXTutor = countTutorias/countHabDirigir
                varEstXTutor = ("{:.2f}".format(varEstXTutor))
                
            ws['K5']=str(varEstXTutor)
            ws['K6']=str(countPrincipales)
            ws['K7']=str(countAsesoresExternos)
            porcentaje = 0 
            if(countPrincipales != 0):
                porcentaje = (countAsesoresExternos/countPrincipales)*100
                porcentaje = ("{:.2f}".format(porcentaje))
            ws['L7']=str(porcentaje)+"%"

            filename = "07__Relacion_Estudiante_Tutor.xlsx"
            #The define the type of response to be give
            response = HttpResponse(content_type = "application/ms-excel")
            content = "attachment; filename = {0}".format(filename)
            response["Content-Disposition"] = content
            wb.save(response)
        
        # -------------------------------- PDF -----------------------------------------

        if(kwargs["type"]==2):#Format pdf request.data["tipo"]==2
            filename = "ReporteInformePorAño.pdf"
            #The define the type of response to be give
            response = HttpResponse(content_type = "application/pdf")
            content = "attachment; filename = {0}".format(filename)
            response["Content-Disposition"] = content
            #Create the PDF object, using the Bytesobject as its "file"
            buffer = BytesIO()
            c = canvas.Canvas(buffer, pagesize=landscape(A4))
            #A4 = (595.275590551181, 841.8897637795275)

            #Header

            c.setLineWidth(.3)
            c.rect(27, 547, 77, 15)
            c.rect(27, 532, 77, 15)
            c.rect(104, 547, 350, 15)
            c.rect(104, 532, 350, 15)
            c.rect(104, 517, 350, 15)
            
            c.setFont('Helvetica-Bold',8)
            c.drawString(30,550,'CARACTERÍSTICA')
            c.drawString(106,550,'7. Relación estudiante / tutor.')

            c.setFont('Helvetica',8)
            c.drawString(67,535,'Indicador')
            c.drawString(106,535,'a. Número de estudiantes por tutor (sólo profesores de TC y habilitados para dirigir tesis).')
            c.drawString(106,520,'c. Número de tutores / asesores externos.')

            high = 475
            wind = 650
            
            c.rect(wind-3, high, 122, 15)
            c.rect(wind-3, high-15, 122, 15)
            c.rect(wind-3, high-30, 122, 15)

            c.setFillColorRGB(0,1,0)
            c.rect(wind+119, high-15, 30, 15, fill=True)
            c.rect(wind+119, high, 30, 15, fill=True)
            c.rect(wind+119, high-30, 30, 15, fill=True)
            c.rect(wind+149, high-30, 30, 15, fill=True)
            
            c.setFillColorRGB(1,0,0)
            c.drawString(wind, high+20,'Periodo: ' + str(InitialPeriod) + ' - ' + str(FinalPeriod))
            c.setFillColorRGB(0,0,0)
            c.drawString(wind, high+3,'Número de estudiantes por tutor:')
            c.drawString(wind,high-12,'Trabajos de Grado:')
            c.drawString(wind,high-26,'Con tutores/asesores externos:')

            #Table header
            styles = getSampleStyleSheet()
            styleBH = styles ["Normal"]
            styleBH.alignment = 1 #TA_CENTER is 1
            styleBH.fontSize=8
            styleBH.Font='Helvetica-Bold'
            high = high - 15
            #						

            profesor = Paragraph('''Profesor''',styleBH)
            habDirigir = Paragraph('''Habilitado para dirigir''',styleBH)
            numTutorias = Paragraph('''Número de Tutorías''',styleBH)
            estudiante = Paragraph('''Estudiante''',styleBH)
            asesorExterno = Paragraph('''Asesor Externo''',styleBH)
            coodirector = Paragraph('''Codirector''',styleBH)
            modalidad = Paragraph('''Modalidad Dirección''',styleBH)
            data = []
            data.append([profesor, habDirigir, numTutorias,estudiante, asesorExterno, coodirector, modalidad])

            #Table
            styleN = styles["BodyText"]
            styleN.alignment = 1 #TA_CENTER is 1
            styleN.fontSize = 8

            countPrincipales = 0
            countAsesoresExternos = 0
            countHabDirigir = 0
            
            countEst = 0
            for q in queryProfessors:
                countEstProf = 0
                professorName = Paragraph(q.user.first_name +" "+ q.user.last_name, styleN)
                
                isDirector = ''
                for q2 in queryDirectors:
                    if(q.id == q2.professor.id):
                        isDirector = Paragraph(str(1), styleN)
                        countHabDirigir += 1
                        break
                profActual = len(data)
                for q2 in queryStudentsProfessor:
                    if (q.user.id == q2.professor.user.id ):
                        countEstProf +=1
                        countEst += 1
                        varRol = ''
                        nameAsesorExterno = ''
                        varModalidad = Paragraph('Co-director', styleN)
                        
                        if(q2.rol == 1):
                            varModalidad = Paragraph('Principal', styleN)
                            countPrincipales +=1
                            for q3 in queryStudentsProfessor:
                                if(q2.student.user.id == q3.student.user.id and q3.rol ==2):
                                    if(q3.professor.is_internal == True):
                                        varRol = Paragraph(q3.professor.user.first_name +" "+ q3.professor.user.last_name, styleN)
                                    else:
                                        nameAsesorExterno = Paragraph(q3.professor.user.first_name +" "+ q3.professor.user.last_name + "(extranjero)", styleN)
                                        countAsesoresExternos+=1
                        estudiante = Paragraph(q2.student.user.first_name +" "+ q2.student.user.last_name, styleN)
                        data.append([professorName, isDirector,'', estudiante, nameAsesorExterno, varRol, varModalidad])
                        high = high - 18
                        professorName = ""
                        isDirector = ""
                
                if(countEstProf == 0):
                    data.append([professorName, isDirector,'0', '', '', '', ''])
                    high = high - 18

                data[profActual][2] = Paragraph(str(countEstProf), styleN)

            varEstXTutor = 0
            porcentaje = 0

            c.drawString(135, high-10,str(countHabDirigir))
            c.drawString(190, high-10,str(countEst))
            if(countHabDirigir != 0):
                varEstXTutor = countEst/countHabDirigir
                varEstXTutor = ("{:.2f}".format(varEstXTutor))
                        
            if(countPrincipales > 0):
                porcentaje = (countAsesoresExternos/countPrincipales) * 100
                porcentaje = ("{:.1f}".format(porcentaje))
            
            c.drawString(777, 478,str(varEstXTutor))
            c.drawString(807, 450,str(porcentaje)+'%')
            c.setFillColorRGB(1,0,0)
            c.drawString(777, 466, str(countPrincipales))
            c.drawString(777, 450,str(countAsesoresExternos))
                
            #Table zise
            width, height = landscape(A4)
            table = Table(data, colWidths=[3*cm, 1.8*cm, 1.9*cm, 4*cm, 4.5*cm,4*cm,2*cm])
            table.setStyle(TableStyle([('INNERGRID',(0,0),(-1,-1),0.25,colors.black),('BOX',(0,0),(-1,-1),0.25,colors.black)]))
            #Pdf size
            table.wrapOn(c, width, height)
            table.drawOn(c,30,high)
            c.showPage()
            
            c.save()
            pdf = buffer.getvalue()
            buffer.close()
            response.write(pdf)
            
        return response

# Modulo B #
def SendEmailNotification(request):
    queryUser = User.objects.get(student=request.data.get('student'))
    queryset = StudentProfessor.objects.filter(student=request.data.get('student')).values('professor')
    for reg in queryset:
        id_Proffesor = reg.get('professor')
        professor = Professor.objects.filter(id=id_Proffesor)[0]
        
        send_email(queryUser, professor)  

class ActivityViewSet(viewsets.ModelViewSet):
    #permission_classes = [IsAuthenticated, (IsStudent | IsDirector | IsCoordinador)]
    queryset = Activity.objects.all()
    serializer_class = ActivitySerializer

class LectureViewSet(viewsets.ModelViewSet):
    #permission_classes = [IsAuthenticated, (IsStudent | IsDirector | IsCoordinador)]
    queryset = Lecture.objects.all()
    serializer_class = LectureSerializer

    def finalize_response(self, request, response, *args, **kwargs):
        response = super().finalize_response(request, response, *args, **kwargs)

        if request.data.get('send_email'):
            SendEmailNotification(request)

        return response

class PublicationViewSet(viewsets.ModelViewSet):
    #permission_classes = [IsAuthenticated, (IsStudent | IsDirector | IsCoordinador)]
    queryset = Publication.objects.all()
    serializer_class = PublicationSerializer

    def finalize_response(self, request, response, *args, **kwargs):
        response = super().finalize_response(request, response, *args, **kwargs)

        if request.data.get('send_email'):
            SendEmailNotification(request)

        return response

class ProjectCourseViewSet(viewsets.ModelViewSet):
    #permission_classes = [IsAuthenticated, (IsStudent | IsDirector | IsCoordinador)]
    queryset = ProjectCourse.objects.all()
    serializer_class = ProjectCourseSerializer

    def finalize_response(self, request, response, *args, **kwargs):
        response = super().finalize_response(request, response, *args, **kwargs)

        if request.data.get('send_email'):
            SendEmailNotification(request)

        return response

class ResearchStaysViewSet(viewsets.ModelViewSet):
    #permission_classes = [IsAuthenticated, (IsStudent | IsDirector | IsCoordinador)]
    queryset = ResearchStays.objects.all()
    serializer_class = ResearchStaysSerializer

    def finalize_response(self, request, response, *args, **kwargs):
        response = super().finalize_response(request, response, *args, **kwargs)

        if request.data.get('send_email'):
            SendEmailNotification(request)

        return response

class PresentationResultsViewSet(viewsets.ModelViewSet):
    #permission_classes = [IsAuthenticated, (IsStudent | IsDirector | IsCoordinador)]
    queryset = PresentationResults.objects.all()
    serializer_class = PresentationResultsSerializer

    def finalize_response(self, request, response, *args, **kwargs):
        response = super().finalize_response(request, response, *args, **kwargs)

        if request.data.get('send_email'):
            SendEmailNotification(request)

        return response

class ParticipationProjectsViewSet(viewsets.ModelViewSet):
    #permission_classes = [IsAuthenticated, (IsStudent | IsDirector | IsCoordinador)]
    queryset = ParticipationProjects.objects.all()
    serializer_class = ParticipationProjectsSerializer

    def finalize_response(self, request, response, *args, **kwargs):
        response = super().finalize_response(request, response, *args, **kwargs)

        if request.data.get('send_email'):
            SendEmailNotification(request)

        return response

class PrizeViewSet(viewsets.ModelViewSet):
    queryset = Prize.objects.all()
    serializer_class = PrizeSerializer

# Consultas a otros modulos #
class ProgramViewSet(viewsets.ModelViewSet):
    queryset = Program.objects.all()
    serializer_class = ProgramSerializer

class InstitutionViewSet(viewsets.ModelViewSet):
    queryset = Institution.objects.all()
    serializer_class = InstitutionSerializer

class InvestigationLineViewSet(viewsets.ModelViewSet):
    queryset = InvestigationLine.objects.all()
    serializer_class = InvestigationLineSerializer

class CityViewSet(viewsets.ModelViewSet):
    queryset = City.objects.all()
    serializer_class = CitySerializer

class CountryViewSet(viewsets.ModelViewSet):
    queryset = Country.objects.all()
    serializer_class = CountrySerializer

# Otro tipo de Consultas #
class TestDirectorAPI(generics.RetrieveAPIView):

    def get(self, request, *args, **kwargs):
        queryset = TestDirector.objects.filter(activity=kwargs['id_activity']).order_by('-date_update')[:1]

        if( len(queryset) > 0 ):
            return Response({"eval_dir": TestDirectorSerializer(queryset[0], many=False).data})
        else:
            return Response({"eval_dir": None})

class TestCoordinatorAPI(generics.RetrieveAPIView):

    def get(self, request, *args, **kwargs):
        queryset = TestCoordinator.objects.filter(activity=kwargs['id_activity']).order_by('-date_update')[:1]

        if( len(queryset) > 0 ):
            return Response({"eval_coord": TestCoordinatorSerializer(queryset[0], many=False).data})
        else:
            return Response({"eval_coord": None})

class PeriodAPI(generics.RetrieveAPIView):

    def get(self, request, *args, **kwargs):
        queryset = Enrrollment.objects.filter(student__user=kwargs['id_user']).values('period').order_by('-period')[:1]

        return HttpResponse(content=json.dumps(queryset[0], cls=DjangoJSONEncoder), status=HTTP_200_OK, content_type="application/json")
        
class PeriodsAPI(generics.RetrieveAPIView):
    serializer_class = PeriodSerializer

    def get(self, request, *args, **kwargs):
        queryset = Enrrollment.objects.filter(student__user=kwargs['id_user']).values('period').order_by('-period').annotate(total=Count('period') )
        return Response({"list_period": PeriodSerializer(queryset, many=True).data})

class ActivitiesAPI(generics.RetrieveAPIView):
    serializer_class = ActivitySerializer

    def get(self, request, *args, **kwargs):
        queryset = Activity.objects.filter(student__user=kwargs['id_user'], academic_year=kwargs['academic_year'], is_active=True)
        return Response({"list_activities": ActivitySerializer(queryset, many=True).data})

class PrizeAPI(generics.RetrieveAPIView):
    serializer_class = PrizeSerializer

    def get(self, request, *args, **kwargs):
        queryset = Prize.objects.filter(activity=kwargs['id_activity'])
        return Response({"prizes": PrizeSerializer(queryset, many=True).data})  

class InvestigatorsAPI(generics.RetrieveAPIView):
    serializer_class = InvestigatorSerializer

    def get(self, request, *args, **kwargs):
        queryset1 = ManageInvestLine.objects.all().values('professor')
        queryinv = Professor.objects.none()
        for reg in queryset1:
            queryinv = queryinv | Professor.objects.filter( id=reg.get('professor') )
        queryset = queryinv
        return Response({"investigadores": InvestigatorSerializer(queryset, many=True).data})  

class InvestigatorAPI(generics.RetrieveAPIView):
    serializer_class = InvestigatorSerializer

    def get(self, request, *args, **kwargs):
        queryset = Professor.objects.filter(id=kwargs['id_professor'])
        return Response({"investigator": InvestigatorSerializer(queryset[0], many=False).data})