from rest_framework import generics, permissions
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from .serializers import UpdateStudentSerializer,StudentSerializer,StudentProfessorSerializer, UpdateGrant, UpdateAgreement, UpdateStudentProfessor
from .models import Student,StudentProfessor, Agreement, Grant, Enrrollment
from d_information_management_app.models import Professor
from d_accounts_app.models import User
from .signals import *

from django.db.models.signals import post_save
from django.db.models import Avg, Sum, Count, Q
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus.paragraph import Paragraph
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle

post_save.connect(change_enrrollment, sender=Enrrollment)
post_save.connect(change_student, sender=Student)

#MIRAR SI ES DE API QUUE DEBE heredar
class ReporteEstudiantesExcel(TemplateView):
    #No se si los kwards se necesitan
    def get(self, request, *args, **kwargs):
        #states={1:"Activo",3:"Graduado",5:"Retirado"}
        states={1:"Activo",2:"Inactivo",3:"Graduado",4:"Balanceado",5:"Retirado"}
        #queryset = Enrrollment.objects.all()
        
        queryset = Enrrollment.objects.filter(Q(state=1) | Q(state=3) | Q(state=5)) 

        if(kwargs["type"]==1):
            wb = Workbook()
            controller = 3
            ws = wb.active
            ws.title = 'Hoja'+str(1)
            
            #Crear titulo
            ws['B1'].alignment = Alignment(horizontal="center",vertical="center")
            ws['B1'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
            top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['B1'].fill = PatternFill(start_color = 'FF0000',end_color ='FFFFFF', fill_type='solid')
            ws['B1'].font = Font(name = 'Calibri',size =16, bold =True)
            ws['B1'] = "Reporte Estudiantes (Activos, Graduados, Retirados) "
            #combinar celdas
            ws.merge_cells('B1:E1')
            ws.row_dimensions[1].height = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
         
            #Crear Cabecera
            ws['B2'].alignment = Alignment(horizontal="center",vertical="center")
            ws['B2'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
            top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['B2'].fill = PatternFill(start_color = '66CFCC',end_color ='66CFCC', fill_type='solid')
            ws['B2'].font = Font(name = 'Calibri',size =12, bold =True)
            ws['B2']='Codigo'

            ws['C2'].alignment = Alignment(horizontal="center",vertical="center")
            ws['C2'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['C2'].fill = PatternFill(start_color = '66CFCC',end_color ='66CFCC', fill_type='solid')
            ws['C2'].font = Font(name = 'Calibri',size =12, bold =True)
            ws['C2']='Nombre'
                    
            ws['D2'].alignment = Alignment(horizontal="center",vertical="center")
            ws['D2'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['D2'].fill = PatternFill(start_color = '66CFCC',end_color ='66CFCC', fill_type='solid')
            ws['D2'].font = Font(name = 'Calibri',size =12, bold =True)
            ws['D2']='Estado'

            ws['E2'].alignment = Alignment(horizontal="center",vertical="center")
            ws['E2'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['E2'].fill = PatternFill(start_color = '66CFCC',end_color ='66CFCC', fill_type='solid')
            ws['E2'].font = Font(name = 'Calibri',size =12, bold =True)
            ws['E2']='Dedicación'
            #Escribir datos en reporte
            for q in queryset:    
                ws.cell(row=controller,column=2).alignment = Alignment(horizontal="center")
                ws.cell(row=controller,column=2).border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                                        top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
                ws.cell(row =controller, column=2).font = Font(name='Calibri',size=12)  
                ws.cell(row =controller, column=4).value = states[q.state]
                ws.cell(row =controller, column=2).value = q.student.user.personal_code

                ws.cell(row=controller,column=3).alignment = Alignment(horizontal="center")
                ws.cell(row=controller,column=3).border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                                        top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
                ws.cell(row =controller, column=3).font = Font(name='Calibri',size=12)       
                ws.cell(row =controller, column=3).value = q.student.user.first_name+" "+q.student.user.last_name 

                ws.cell(row=controller,column=4).alignment = Alignment(horizontal="center")
                ws.cell(row=controller,column=4).border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                                        top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
                ws.cell(row =controller, column=4).font = Font(name='Calibri',size=12)
                ws.cell(row =controller, column=4).value = states[q.state]

                ws.cell(row=controller,column=5).alignment = Alignment(horizontal="center")
                ws.cell(row=controller,column=5).border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                                        top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
                ws.cell(row =controller, column=5).font = Font(name='Calibri',size=12)       
                ws.cell(row =controller, column=5).value = "Completo" if q.student.dedication==1 else "Parcial" 
                controller +=1    

            filename = "ReporteEstudiantes.xlsx"                             
            response = HttpResponse(content_type = "application/ms-excel")
            content = "attachment; filename = {0}".format(filename)
            response["Content-Disposition"] = content
            wb.save(response)
            
        if(kwargs["type"]==2):
            filename = "ReporteEstudiantes.pdf"
           
            response = HttpResponse(content_type = "application/pdf")
            content = "attachment; filename = {0}".format(filename)
            response["Content-Disposition"] = content
            #Create the PDF object, using the Bytesobject as its "file"
            buffer = BytesIO()
            c = canvas.Canvas(buffer, pagesize=A4)

            c.setLineWidth(.3)
            c.setFont('Helvetica',22)
            c.drawString(30,750,'Reporte Estudiantes (Activos, Graduados, Retirados)')
            c.setFont('Helvetica',14)
            c.setFont('Helvetica-Bold',12)
            c.line(460,747,560,747)
            #CABECERA
            styles = getSampleStyleSheet()
            styleBH = styles ["Normal"]
            styleBH.alignment = 1 
            styleBH.fontSize=10

            codigo = Paragraph('''Codigo''',styleBH)
            nombre = Paragraph('''Nombre''',styleBH)
            estado = Paragraph('''Estado''',styleBH)
            dedicacion = Paragraph('''Dedicación''',styleBH)
            data = []
            data.append([codigo, nombre, estado, dedicacion])

            #TABLA
            styleN = styles["BodyText"]
            styleN.alignment = 1 
            styleN.fontSize = 7
            high = 650
            for q in queryset: 
                data.append([ q.student.user.personal_code,
                              q.student.user.first_name+" "+q.student.user.last_name,
                              states[q.state],
                              "Completo" if q.student.dedication==1 else "Parcial"])
                high = high -18
            #Table zise
            width, height = A4
            table = Table(data, colWidths=[2.7*cm,6.9*cm,2.2*cm,2.2*cm,2.2*cm,2.7*cm])
            table.setStyle(TableStyle([('INNERGRID',(0,0),(-1,-1),0.25,colors.black),('BOX',(0,0),(-1,-1),0.25,colors.black)]))
            #Pdf size
            table.wrapOn(c, width, height)
            table.drawOn(c,30,high)
            c.showPage()
            
            c.save()
            pdf = buffer.getvalue()
            buffer.close()
            response.write(pdf)           

        return response


class CreateStudentProfessor(generics.GenericAPIView):
    serializer_class=StudentProfessorSerializer
    def post (self,request):
        
        serializer= self.get_serializer(data=request.data)
        if serializer.is_valid():
            
            '''is_element = StudentProfessor.objects.filter(student=request.data["student"],professor=request.data["professor"])
            if not (is_element):'''

            if (int(request.data["rol"])==1):
                
                isdirector= StudentProfessor.objects.filter(student=request.data["student"],
                professor=request.data["professor"], rol= request.data["rol"])
            
                if not(isdirector):
                    serializer.save()
                    return Response(serializer.data,status=status.HTTP_201_CREATED)
                return Response(f"Ya cuenta con su director",status=status.HTTP_400_BAD_REQUEST)
            else:
                
                iscodirector= StudentProfessor.objects.filter(student=request.data["student"], rol= request.data["rol"]) 
                if (len(iscodirector)<2):
                    serializer.save()
                    return Response(serializer.data,status=status.HTTP_201_CREATED)
                return Response(f"Ya cuenta con dos coodirectores",status=status.HTTP_400_BAD_REQUEST)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)            
            

            
        


class UpdateStudenAPI (APIView):
    def get (self,request,*args,**kwargs):
        model=Student.objects.filter(id=kwargs["id"])
        return Response(UpdateStudentSerializer(model, many=True).data,status=status.HTTP_202_ACCEPTED)
    def put (self,request,*args,**kwargs):
        try:
            model = Student.objects.get(id=kwargs["id"])
        except Student.DoesNotExist:
            return Response(f"El Estudiante con el ID solicitado no existe en la Base de Datos.",status= status.HTTP_404_NOT_FOUND)
        serializer = UpdateStudentSerializer(model, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_202_ACCEPTED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class UpdateGrantAPI (APIView):
    def get (self,request,*args,**kwargs):
        model= Grant.objects.filter(id=kwargs["id"])
        return Response(UpdateGrant(model, many=True).data,status=status.HTTP_202_ACCEPTED)
    def put (self,request,*args,**kwargs):
        try:
            model = Grant.objects.get(id=kwargs["id"])
        except Grant.DoesNotExist:
            return Response(f"La beca solicitada no existe en la Base de Datos.",status= status.HTTP_404_NOT_FOUND)
        serializer = UpdateGrant(model, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_202_ACCEPTED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class UpdateAgreementAPI (APIView):
    def get (self,request,*args,**kwargs):
        model= Agreement.objects.filter(id=kwargs["id"])
        return Response(UpdateAgreement(model, many=True).data,status=status.HTTP_202_ACCEPTED)
    def put (self,request,*args,**kwargs):
        try:
            model = Agreement.objects.get(id=kwargs["id"])
        except Agreement.DoesNotExist:
            return Response(f"El convenio solicitado no existe en la Base de Datos.",status= status.HTTP_404_NOT_FOUND)
        serializer = UpdateAgreement(model, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_202_ACCEPTED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class UpdateStudentProfessorAPI (APIView):
    def get (self,request,*args,**kwargs):
        model= StudentProfessor.objects.filter(student=kwargs["id_s"],professor=kwargs["id_p"])
        return Response(UpdateStudentProfessor(model, many=True).data,status=status.HTTP_202_ACCEPTED)
    def put (self,request,*args,**kwargs):
        try:
            model = StudentProfessor.objects.get(student=kwargs["id_s"],professor=kwargs["id_p"])
        except StudentProfessor.DoesNotExist:
            return Response(f"Los codirectores y director solicitados no existe en la Base de Datos.",status= status.HTTP_404_NOT_FOUND)
        serializer = UpdateStudentProfessor(model, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_202_ACCEPTED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



