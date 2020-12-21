#Este va servir como un estilo de view.py, este es el enlace el cual recibe la peticion y
#se comunica con el serializador

import datetime

from io import BytesIO
from datetime import datetime

from django.db.models.signals import post_save
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView

from rest_framework import generics, permissions, status, views, viewsets
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.response import Response
from rest_framework.views import APIView

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
from reportlab.platypus.paragraph import Paragraph

from a_students_app.models import StudentGroupInvestigation, Student, Enrrollment
from d_accounts_app.backends import IsProfessor, IsCoordinator
from d_accounts_app.models import User

from .models import *
from .serializers import *
from .signals import *

post_save.connect(change_professor, sender=Professor)
post_save.connect(change_country, sender=Country)
post_save.connect(change_state, sender=State)
post_save.connect(change_city, sender=City)
post_save.connect(change_inv_group, sender=InvestigationGroup)
post_save.connect(change_know_area, sender=KnowledgeArea)
post_save.connect(change_inv_line, sender=InvestigationLine)
post_save.connect(change_user, sender=User)

# Create your api's here.
# --------------------------------------------------Arias

#region Create
class ReportTest(APIView):
    def get(self, request, *args, **kwargs):

        states={1:"Activo",2:"Inactivo",3:"Graduado",4:"Balanceado",5:"Retirado"}
           
        year = kwargs["year"]
        queryset = Enrrollment.objects.filter(admission_date__range=(str(year)+"-1-1",str(year)+"-12-31"))
        now = datetime.datetime.now()

        if(kwargs["type"]==1):#Format xlsx request.data["tipo"]==1
            wb = Workbook()
            controller = 3
            ws = wb.active
            ws.title = 'Hoja'+str(1)

            #Create title in the sheet
            ws['B1'].alignment = Alignment(horizontal="center",vertical="center")
            ws['B1'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
            top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['B1'].fill = PatternFill(start_color = '66FFCC',end_color ='66FFCC', fill_type='solid')
            ws['B1'].font = Font(name = 'Calibri',size =14, bold =True)
            ws['B1'] = "INFORME DEL AÑO "+str(year)
            #Change the characteristics of cells
            ws.merge_cells('B1:G1')
            ws.row_dimensions[1].height = 25
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 25
            ws.column_dimensions['G'].width = 20
            #Create header
            ws['B2'].alignment = Alignment(horizontal="center",vertical="center")
            ws['B2'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
            top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['B2'].fill = PatternFill(start_color = '66CFCC',end_color ='66CFCC', fill_type='solid')
            ws['B2'].font = Font(name = 'Calibri',size =12, bold =True)
            ws['B2']='Codigo'

            ws['C2'].alignment = Alignment(horizontal="center",vertical="center")
            ws['C2'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['C2'].fill = PatternFill(start_color = '66CFCC',end_color ='66CFCC', fill_type='solid')
            ws['C2'].font = Font(name = 'Calibri',size =12, bold =True)
            ws['C2']='Nombre'
                    
            ws['D2'].alignment = Alignment(horizontal="center",vertical="center")
            ws['D2'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['D2'].fill = PatternFill(start_color = '66CFCC',end_color ='66CFCC', fill_type='solid')
            ws['D2'].font = Font(name = 'Calibri',size =12, bold =True)
            ws['D2']='Estado'

            ws['E2'].alignment = Alignment(horizontal="center",vertical="center")
            ws['E2'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['E2'].fill = PatternFill(start_color = '66CFCC',end_color ='66CFCC', fill_type='solid')
            ws['E2'].font = Font(name = 'Calibri',size =12, bold =True)
            ws['E2']='Dedicación'

            ws['F2'].alignment = Alignment(horizontal="center",vertical="center")
            ws['F2'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['F2'].fill = PatternFill(start_color = '66CFCC',end_color ='66CFCC', fill_type='solid')
            ws['F2'].font = Font(name = 'Calibri',size =12, bold =True)
            ws['F2']='Institución'

            ws['G2'].alignment = Alignment(horizontal="center",vertical="center")
            ws['G2'].border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
            ws['G2'].fill = PatternFill(start_color = '66CFCC',end_color ='66CFCC', fill_type='solid')
            ws['G2'].font = Font(name = 'Calibri',size =12, bold =True)
            ws['G2']='Departamento'

            for q in queryset:    
                ws.cell(row=controller,column=2).alignment = Alignment(horizontal="center")
                ws.cell(row=controller,column=2).border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                                        top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
                ws.cell(row =controller, column=2).font = Font(name='Calibri',size=12)       
                ws.cell(row =controller, column=2).value = q.student.user.personal_code

                ws.cell(row=controller,column=3).alignment = Alignment(horizontal="center")
                ws.cell(row=controller,column=3).border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                                        top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
                ws.cell(row =controller, column=3).font = Font(name='Calibri',size=12)       
                ws.cell(row =controller, column=3).value = q.student.user.first_name+" "+q.student.user.last_name 

                ws.cell(row=controller,column=4).alignment = Alignment(horizontal="center")
                ws.cell(row=controller,column=4).border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                                        top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
                ws.cell(row =controller, column=4).font = Font(name='Calibri',size=12)
                ws.cell(row =controller, column=4).value = states[q.state]

                ws.cell(row=controller,column=5).alignment = Alignment(horizontal="center")
                ws.cell(row=controller,column=5).border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                                        top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
                ws.cell(row =controller, column=5).font = Font(name='Calibri',size=12)       
                ws.cell(row =controller, column=5).value = "Completo" if q.student.dedication==1 else "Parcial" 
                
                ws.cell(row=controller,column=6).alignment = Alignment(horizontal="center")
                ws.cell(row=controller,column=6).border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                                        top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
                ws.cell(row =controller, column=6).font = Font(name='Calibri',size=12)       
                ws.cell(row =controller, column=6).value = q.student.program.deparment.faculty.institution.name_inst

                ws.cell(row=controller,column=7).alignment = Alignment(horizontal="center")
                ws.cell(row=controller,column=7).border = Border(left = Side(border_style = "thin"), right=Side(border_style = "thin"),
                                                                        top=Side(border_style="thin"),bottom=Side(border_style ="thin"))
                ws.cell(row =controller, column=7).font = Font(name='Calibri',size=12)       
                ws.cell(row =controller, column=7).value = q.student.program.deparment.name 
                controller +=1

            filename = "ReporteInformePorAño.xlsx"
            #The define the type of response to be give
            response = HttpResponse(content_type = "application/ms-excel")
            content = "attachment; filename = {0}".format(filename)
            response["Content-Disposition"] = content
            wb.save(response)

        if(kwargs["type"]==2):#Format pdf request.data["tipo"]==2
            filename = "ReporteInformePorAño.pdf"
            #The define the type of response to be give
            response = HttpResponse(content_type = "application/pdf")
            content = "attachment; filename = {0}".format(filename)
            response["Content-Disposition"] = content
            #Create the PDF object, using the Bytesobject as its "file"
            buffer = BytesIO()
            c = canvas.Canvas(buffer, pagesize=A4)


            #Header
            c.setLineWidth(.3)
            c.setFont('Helvetica',22)
            c.drawString(30,750,'Reporte anual')
            c.setFont('Helvetica',12)
            c.drawString(30,735,str(now.year)+"-"+str(now.month)+"-"+str(now.day))
            c.setFont('Helvetica-Bold',12)
            c.drawString(460,750,str(year))
            c.line(460,747,560,747)

            #Table header
            styles = getSampleStyleSheet()
            styleBH = styles ["Normal"]
            styleBH.alignment = 1 #TA_CENTER is 1
            styleBH.fontSize=10

            codigo = Paragraph('''Codigo''',styleBH)
            nombre = Paragraph('''Nombre''',styleBH)
            estado = Paragraph('''Estado''',styleBH)
            dedicacion = Paragraph('''Dedicación''',styleBH)
            institucion = Paragraph('''Institución''',styleBH)
            departamento = Paragraph('''Departamento''',styleBH)
            data = []
            data.append([codigo, nombre, estado, dedicacion, institucion, departamento])

            #Table
            styleN = styles["BodyText"]
            styleN.alignment = 1 #TA_CENTER is 1
            styleN.fontSize = 7

            high = 650
            for q in queryset: 
                data.append([ q.student.user.personal_code,
                              q.student.user.first_name+" "+q.student.user.last_name,
                              states[q.state],
                              "Completo" if q.student.dedication==1 else "Parcial",
                              q.student.program.deparment.faculty.institution.name_inst,
                              q.student.program.deparment.name])
                high = high -18
            
            #Table zise
            width, height = A4
            table = Table(data, colWidths=[2.7*cm,5.4*cm,2.2*cm,2.2*cm,4.0*cm,2.7*cm])
            table.setStyle(TableStyle([('INNERGRID',(0,0),(-1,-1),0.25,colors.black),('BOX',(0,0),(-1,-1),0.25,colors.black)]))
            #Pdf size
            table.wrapOn(c, width, height)
            table.drawOn(c,30,high)
            c.showPage()
            
            c.save()
            pdf = buffer.getvalue()
            buffer.close()
            response.write(pdf)           

        return response        

class CountryAPI(viewsets.ModelViewSet):
    """
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    API que permite:
    ☠ Crear un País, esta función hace uso del metodo POST.
    PATH: 'api/1.0/crear_pais/'
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    """
    #permission_classes = [IsAuthenticated, IsCoordinator]
    serializer_class = CountrySerializer
    queryset = Country.objects.all()


class StateAPI(viewsets.ModelViewSet):
    """
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒►Departamento en contexto de País
    API que permite:
    ☠ Crear un Departamento, esta función hace uso del metodo POST.
    PATH: '/api/state/'
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    """
    #permission_classes = [IsAuthenticated, IsCoordinator]
    serializer_class = StateSerializer
    queryset = State.objects.all()

class CityAPI(viewsets.ModelViewSet):
    """
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    API que permite:
    ☠ Crear una Ciudad, esta función hace uso del metodo POST.
    PATH: '/api/city/'
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    """
    #permission_classes = [IsAuthenticated, IsCoordinator]
    serializer_class = CitySerializer
    queryset = City.objects.all()

class InstitutionAPI(viewsets.ModelViewSet):
    """
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    API que permite:
    ☠ Crear una Institución, esta función hace uso del metodo POST.
    PATH: '/api/Institution/'
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    """
    #permission_classes = [IsAuthenticated, IsCoordinator]
    serializer_class = InstitutionSerializer
    queryset = Institution.objects.all()

class ProfessorAPI(viewsets.ModelViewSet):
    """
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    API que permite:
    ☠ Crear un Profesor, esta api hace uso del metodo POST.
    PATH: 'api/1.0/crear_profesor/'
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    """
    #permission_classes = [IsAuthenticated, IsCoordinator]
    serializer_class = ProfessorSerializer
    queryset = Professor.objects.all()

class FacultyAPI(viewsets.ModelViewSet):
    """
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    API que permite:
    ☠ Crear una Facultad, esta función hace uso del metodo POST.
    PATH: '/api/faculty/'
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    """
    #permission_classes = [IsAuthenticated, IsCoordinator]
    serializer_class = FacultySerializer
    queryset = Faculty.objects.all()

class DepartmentAPI(viewsets.ModelViewSet):
    """
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒►Departamento en contexto de Facultad
    API que permite:
    ☠ Crear un Departamento, esta función hace uso del metodo POST.
    PATH: '/api/department_u/'
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    """
    #permission_classes = [IsAuthenticated, IsCoordinator]
    serializer_class = DepartmentSerializer
    queryset = Department.objects.all()
#endregion

#region Consult

class ConsultCountryAPI(APIView):
    """
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    API que permite:
    ☠ Consultar Paises, esta función hace uso del metodo GET. 
    PATH: 'api/1.0/consultar_pais/'
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    """
    #permission_classes = [IsAuthenticated, IsCoordinator]
    def get(self, request, *args, **kwargs):
        queryset = Country.objects.filter(status=True)
        return Response({"Countrys": CountrySerializer(queryset, many=True).data })  

class ConsultState_CountryAPI(APIView):
    """
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒►Departamento en contexto de País
    API que permite:
    ☠ Consultar Departamentos de un determinado País enviando su id, esta función hace uso del metodo GET.
    ☠ Actualizar un País enviando un JSON, esta función hace uso del método PUT.
    PATH: 'api/1.0/consultar_departamento_pais/<int:id_country>'
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    """
    #permission_classes = [IsAuthenticated, IsCoordinator]
    def get(self, request, *args, **kwargs):
        queryset = State.objects.filter(country=kwargs["id_country"], status=True)
        returned = StateSerializer(queryset, many=True).data
        if returned:
            return Response({"States": returned}, status=status.HTTP_202_ACCEPTED)
        else:
            return Response(f"No existen Departamentos con ese País en la base de datos", status=status.HTTP_404_NOT_FOUND)

class FullConsultState_CountryAPI(APIView):
    """
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒►Departamento en contexto de País
    API que permite:
    ☠ Consultar Departamentos (sin filtrar por status) de un determinado País enviando su id, esta función hace uso del metodo GET.
    PATH: 'api/1.0/consultar_departamento_pais/<int:id_country>'
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    """
    #permission_classes = [IsAuthenticated, IsCoordinator]
    def get(self, request, *args, **kwargs):
        try:
            country = Country.objects.get(id=kwargs["id_country"])
        except Country.DoesNotExist:
            return Response(f"No existe País en la base de datos", status=status.HTTP_404_NOT_FOUND)

        queryset = State.objects.filter(country=country)
        returned = StateSerializer(queryset, many=True).data
        if returned:
            return Response({"States": returned}, status=status.HTTP_202_ACCEPTED)
        else:
            return Response(f"No existen Departamentos con ese País en la base de datos", status=status.HTTP_404_NOT_FOUND)

class ConsultCity_StateAPI(APIView):
    """
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    API que permite:
    ☠ Consultar Ciudades de un determinado Departamento enviando su id, esta función hace uso del metodo GET.
    ☠ Actualizar un Departamento enviando un JSON, esta función hace uso del método PUT.
    PATH: 'api/1.0/consultar_ciudad_departamento/<int:id_dep>'
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    """
    #permission_classes = [IsAuthenticated, IsCoordinator]
    def get(self, request, *args, **kwargs):
        queryset = City.objects.filter(state=kwargs["id_dep"], status=True)
        returned = CitySerializer(queryset, many=True).data
        if returned:
            return Response({"Cities": returned}, status=status.HTTP_202_ACCEPTED)
        else:
            return Response(f"No existen Ciudades con ese Departamento en la base de datos", status=status.HTTP_404_NOT_FOUND)

    def put(self, request, *args, **kwargs):
        try:
            model = City.objects.get(id=request.data['id'])
        except City.DoesNotExist:
            return Response(f"No existe esa Ciudad en la base de datos", status=status.HTTP_404_NOT_FOUND)
        
        serializer = CitySerializer(model, data=request.data)
        if serializer.is_valid():
            if 'status' in request.data.keys():
                if request.data['status'] == False:
                    institutions = Institution.objects.filter(city=model)
                    for regI in institutions:
                        tmpInstitution = Institution.objects.get(id=regI.id)
                        tmpInstitution.status = False
                        tmpInstitution.save()
                if request.data['status'] == True:
                    institutions = Institution.objects.filter(city=model)
                    for regI in institutions:
                        tmpInstitution = Institution.objects.get(id=regI.id)
                        tmpInstitution.status = True
                        tmpInstitution.save()
            serializer.save()
            return Response(serializer.data, status=status.HTTP_202_ACCEPTED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class FullConsultCity_StateAPI(APIView):
    """
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    API que permite:
    ☠ Consultar Ciudades (sin filtrar por status) de un determinado Departamento enviando su id, esta función hace uso del metodo GET.
    PATH: 'api/1.0/consultar_ciudad_departamento/<int:id_dep>'
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    """
    #permission_classes = [IsAuthenticated, IsCoordinator]
    def get(self, request, *args, **kwargs):
        try:
            state = State.objects.get(id=kwargs["id_dep"])
        except State.DoesNotExist:
            return Response(f"No existe Departamento en la base de datos", status=status.HTTP_404_NOT_FOUND)

        queryset = City.objects.filter(state=state)
        returned = CitySerializer(queryset, many=True).data
        if returned:
            return Response({"Cities": returned}, status=status.HTTP_202_ACCEPTED)
        else:
            return Response(f"No existen Ciudades con ese Departamento en la base de datos", status=status.HTTP_404_NOT_FOUND)

class ConsultInstitutionAPI(APIView):
    """
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    API que permite:
    ☠ Consultar Instituciones, esta función hace uso del metodo GET.
    PATH: 'api/1.0/consultar_institucion/'
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    """
    def get(self, request, *args, **kwargs):
        queryset = Institution.objects.filter(status=True)
        return Response({"Institutions": InstitutionSerializer(queryset, many=True).data })

class ConsultFacultyAPI(APIView):
    """
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    API que permite:
    ☠ Consultar Facultades, esta función hace uso del metodo GET.
    PATH: 'api/1.0/consultar_facultad/'
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    """
    def get(self, request, *args, **kwargs):
        queryset = Faculty.objects.filter(status=True)
        return Response({"Facultys": FacultySerializer(queryset, many=True).data })

class ConsultDepartmentAPI(APIView):
    """
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒►Departamento en contexto de Facultad
    API que permite:
    ☠ Consultar Departamentos, esta función hace uso del metodo GET.
    PATH: 'api/1.0/consultar_departamentoU/'
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    """
    def get(self, request, *args, **kwargs):
        queryset = Department.objects.filter(status=True)
        return Response({"Departments": DepartmentSerializer(queryset, many=True).data })

class FullConsultInstitution_CityAPI(APIView):
    """
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    API que permite:
    ☠ Consultar Instituciones por ciudad, esta función hace uso del metodo GET.
    PATH: 'api/1.0/full_consultar_institucion_ciudad/<int:id_city>'
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    """
    def get(self, request, *args, **kwargs):
        try:
            City.objects.get(id=kwargs["id_city"])
        except City.DoesNotExist:
            return Response(f"No existe usa ciudad en la base de datos para los datos ingresados", status=status.HTTP_404_NOT_FOUND)
                    
        queryset = Institution.objects.filter(city=kwargs["id_city"])
        return Response({"Institutions": InstitutionSerializer(queryset, many=True).data }, status=status.HTTP_202_ACCEPTED)

#endregion

# Create your api's here.
# --------------------------------------------------Jeison

#region Create
# Grupo de investigacion
class InvestigationGroupAPI(viewsets.ModelViewSet):
    """
    Clase usada para la implementacion de la API para crear un 
    Grupo de Investigacion
    """
    #permission_classes = [IsAuthenticated, IsCoordinator]
    serializer_class = InvestigationGroupSerializer
    queryset = InvestigationGroup.objects.all()

# Area del conocimiento
class KnowledgeAreaAPI(viewsets.ModelViewSet):
    """
    Clase usada para la implementacion de la API para crear un 
    Area del Conocimiento
    """
    #permission_classes = [IsAuthenticated, IsCoordinator]
    serializer_class = KnowledgeAreaSerializer
    queryset = KnowledgeArea.objects.all()

# Linea de investigacion
class InvestigationLineAPI(viewsets.ModelViewSet):
    """
    Clase usada para la implementacion de la API para crear una
    Linea de Investigacion
    """
    #permission_classes = [IsAuthenticated, IsCoordinator]
    serializer_class = InvestigationLineSerializer
    queryset = InvestigationLine.objects.all()

# Formacion academica
class AcademicTrainingAPI(viewsets.ModelViewSet):
    """
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    API que permite:
    ☠ Crear una Formación Academica, esta función hace uso del metodo POST.
    PATH: '/api/academic_training/'
    ▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒
    """
    #permission_classes = [IsAuthenticated, IsCoordinator]
    serializer_class = AcademicTrainingSerializer
    queryset = AcademicTraining.objects.all()

# Trabaja entre GI y AC
class CreateWorksInvestGroupAPI(generics.GenericAPIView):
    """
    Clase usada para la implementacion de la API para crear una relacion
    de rol Trabaja entre los modelos de Grupo de Investigacion y Area del Conocimiento
    """
    serializer_class = WorksInvestGroupSerializer

    def post(self, request):
        serializer = self.get_serializer(data = request.data)
        if serializer.is_valid():
            korksInvestGroup = serializer.save()
            return Response(serializer.data, status = status.HTTP_201_CREATED)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

# Dirige entre P y GI
class CreateManageInvestGroupAPI(generics.GenericAPIView):
    """
    Clase usada para la implementacion de la API para crear una relacion
    de rol Dirige entre los modelos de Profesor y Grupo de Investigacion
    """
    serializer_class = ManageInvestGroupSerializer

    def post(self, request):
        serializer = self.get_serializer(data = request.data)
        if serializer.is_valid():
            directs = serializer.save()
            return Response(serializer.data, status = status.HTTP_201_CREATED)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

# Es miembro entre P y GI
class CreateIsMemberAPI(generics.GenericAPIView):
    """
    Clase usada para la implementacion de la API para crear una relacion
    de rol Es Miembro entre los modelos de Profesor y Grupo de Investigacion
    """
    serializer_class = IsMemberSerializer

    def post(self, request):
        serializer = self.get_serializer(data = request.data)
        if serializer.is_valid():
            is_member = serializer.save()
            return Response(serializer.data, status = status.HTTP_201_CREATED)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

# Maneja entre P y LI
class CreateManageInvestLineAPI(generics.GenericAPIView):
    """
    Clase usada para la implementacion de la API para crear una relacion
    de rol Maneja entre los modelos de Profesor y Linea de Investigacion
    """
    serializer_class = ManageInvestLineSerializer

    def post(self, request):
        serializer = self.get_serializer(data = request.data)
        if serializer.is_valid():
            drive = serializer.save()
            return Response(serializer.data, status = status.HTTP_201_CREATED)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

# Labora entre P y Departamento de la U
class CreateWorkDepartmentAPI(generics.GenericAPIView):
    """
    Clase usada para la implementacion de la API para crear una relacion
    de rol Labora entre los modelos de Profesor y Departamento (de la Universidad)
    """
    serializer_class = WorksDepartmSerializer

    def post(self, request):
        serializer = self.get_serializer(data = request.data)
        if serializer.is_valid():
            work = serializer.save()
            return Response(serializer.data, status = status.HTTP_201_CREATED)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

# Director entre profesor y programa
class CreateCoordinatorProgramAPI(generics.GenericAPIView):
    """
    Clase usada para la implementacion de la API para crear una relacion
    de rol coordinador entre los modelos de Profesor y Programa (de la Universidad)
    """
    serializer_class = CoordinatorProgramSerializer

    def post(self, request):
        serializer = self.get_serializer(data = request.data)
        if serializer.is_valid():
            coordinator = serializer.save()
            return Response(serializer.data, status = status.HTTP_201_CREATED)
        return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

#endregion

#region Consult
class ConsultInvestigationGroup_DepartmentAPI(APIView):
    """
    Clase usada para la implementacion de la API para consultar todos los
    Grupos de Investigacion que pertenecen a un Departamento espesifico de la Universidad,
    esto se logra enviando el ID del departamento mediante el metodo GET
    - - - - -
    Parameter
    - - - - -
    dep : int
        Referencia a un departamento
    """
    #permission_classes = [IsAuthenticated, IsCoordinator]
    def get(self, request, *args, **kwargs):
        queryset = InvestigationGroup.objects.filter(department=kwargs['dep'], status=True)
        return Response({"Groups": InvestigationGroupSerializer(queryset, many=True).data }, status = status.HTTP_200_OK)

class ConsultInvestigationGroup_InstAPI(APIView):
    """
    Clase usada para la implementacion de la API para consultar todos los
    Grupos de Investigacion que pertenecen a un Departamento espesifico de la Universidad,
    esto se logra enviando el ID del departamento mediante el metodo GET
    - - - - -
    Parameter
    - - - - -
    id : int
        Referencia a una institucion
    """
    #permission_classes = [IsAuthenticated, IsCoordinator]
    def get(self, request, *args, **kwargs):
        queryset = InvestigationGroup.objects.filter(department__faculty__institution=kwargs['id'], status=True)
        return Response({"Groups": InvestigationGroupSerializer(queryset, many=True).data }, status = status.HTTP_200_OK)

#profesores
class ConsultProfessorAPI(APIView):
    """
    Clase usada para la implementacion de la API para consultar todos los Profesores
    registrados en la Universidad
    """
    #permission_classes = [IsAuthenticated, IsCoordinator]
    def get(self, request, *args, **kwargs):
        queryset = Professor.objects.filter(status=True)
        return Response({"Professors": ProfessorSerializer(queryset, many=True).data }, status = status.HTTP_200_OK)

class ConsultProfessor_userAPI(APIView):
    """
    Clase usada para la implementacion de la API para consultar un Profesor espesifico
    de la Universidad, esto se logra enviando el ID del Usuario asigano al Profesor mediante 
    el metodo GET
    - - - - -
    Parameters
    - - - - -
    id : int
        Referencia a un Usuario
    """
    #permission_classes = [IsAuthenticated, IsCoordinator]
    def get(self, request, *args, **kwargs):
        queryset = Professor.objects.filter(user=kwargs['id'], status=True)

        returned = ProfessorSerializer(queryset, many=True).data
        if returned:
            return Response({"Professor": returned}, status=status.HTTP_202_ACCEPTED)
        else:
            return Response(f"No existe el Profesor en la base de datos", status=status.HTTP_404_NOT_FOUND)

class ConsultProfessorDirectorGIAPI(APIView):
    def get(self, request, *args, **kwargs):
        filterProf = WorksDepartm.objects.filter(laboral_category="planta", professor__is_internal=True)
        queryset = []
        for ref in filterProf:
            aux = Professor.objects.filter(id=ref.professor.id)
            if aux:
                queryset.extend(aux)
        return Response({"Professors": ProfessorSerializer(queryset, many=True).data }, status = status.HTTP_200_OK)

class ConsultProfessorDirectorStudentAPI(APIView):
    def get(self, request, *args, **kwargs):
        filterProf = WorksDepartm.objects.filter(laboral_category="planta", professor__is_internal=True)
        filterProf2 = WorksDepartm.objects.filter(laboral_category="catedra", professor__is_internal=True)
        queryset = []
        for ref in filterProf:
            aux = Professor.objects.filter(id=ref.professor.id)
            if aux:
                queryset.extend(aux)
        for ref in filterProf2:
            aux = Professor.objects.filter(id=ref.professor.id)
            if aux:
                queryset.extend(aux)
        return Response({"Professors": ProfessorSerializer(queryset, many=True).data }, status = status.HTTP_200_OK)

class ConsultProfessorCoodirectorPlantaAPI(APIView):
    def get(self, request, *args, **kwargs):
        filterProf = WorksDepartm.objects.filter(laboral_category="planta")
        queryset = []
        for ref in filterProf:
            aux = Professor.objects.filter(id=ref.professor.id)
            if aux:
                queryset.extend(aux)
        return Response({"Professors": ProfessorSerializer(queryset, many=True).data }, status = status.HTTP_200_OK)

# Area del conocimiento
class ConsultKnowledgeAreaAPI(APIView):
    """
    Clase usada para la implementacion de la API para consultar todas las Areas del Conocimiento 
    registradas en la Universidad
    """
    #permission_classes = [IsAuthenticated, IsCoordinator]
    def get(self, request, *args, **kwargs):
        queryset = KnowledgeArea.objects.filter(status=True)
        return Response({"Knowledges": KnowledgeAreaSerializer(queryset, many=True).data }, status = status.HTTP_200_OK)

# Lineas de investigacion
class ConsultInvestigationLine_knowledgeAPI(APIView):
    """
    Clase usada para la implementacion de la API para consultar todas las Lineas de Investigacion que
    pertenecen a un Area del conocimiento espesifica
    """
    #permission_classes = [IsAuthenticated, IsCoordinator]
    def get(self, request, *args, **kwargs):
        queryset = InvestigationLine.objects.filter(know_area=kwargs['id_area'], status=True)
        returned = InvestigationLineSerializer(queryset, many=True).data
        if returned:
            return Response({"Lines": InvestigationLineSerializer(queryset, many=True).data }, status = status.HTTP_200_OK)
        else:
            return Response(f"No existen Lineas de Investigacion asociadas a esa Area del conocimiento...", status = status.HTTP_400_BAD_REQUEST)

class ConsultInvestigationLine_GIAPI(APIView):
    """
    Clase usada para la implementacion de la API para consultar todas las Lineas de Investigacion que
    pertenecen a un Area del conocimiento espesifica
    """
    #permission_classes = [IsAuthenticated, IsCoordinator]
    def get(self, request, *args, **kwargs):
        know_filter = WorksInvestGroup.objects.filter(inv_group=kwargs['id_gi'])
        queryset = []
        for work in know_filter:
            aux = InvestigationLine.objects.filter(know_area=work.know_area, status=True)
            if aux:
                queryset.extend(aux)
        returned = InvestigationLineSerializer(queryset, many=True).data
        if returned:
            return Response({"Lines": InvestigationLineSerializer(queryset, many=True).data }, status=status.HTTP_200_OK)
        else:
            return Response(f"No existen Lineas de Investigacion asociadas a ese grupo de investigacion...", status=status.HTTP_400_BAD_REQUEST)

# Es miembro
class ConsultIsMemberAPI(APIView):
    """
    Clase usada para la implementacion de, la API para consultar si un profesor ES MIEMBRO o no de un 
    grupo de investigacion espesifico de la Universidad, esto se logra enviando el ID del Profesor y 
    el ID del Grupo de Investigacion (en ese orden) mediante el metodo GET y/o enviando la informacion 
    que se va a editar del registro del modelo "Es Miembro" mediante el metodo PUT
    - - - - -
    Parameters
    - - - - -
    id_p : int
        Referencia a un Profesor
    id_gi : int
        Referencia a un Grupo de Investigacion
    """
    def get(self, request, *args, **kwargs):
        queryset = IsMember.objects.filter(inv_group=kwargs['id_gi'], professor=kwargs['id_p'], member_status=True)
        returned = IsMemberSerializer(queryset, many=True).data
        if returned:
            return Response({"IsMember":returned}, status=status.HTTP_202_ACCEPTED)
        else:
            return Response(f"No existe un registro en la base de datos para los datos ingresados", status=status.HTTP_404_NOT_FOUND)
    
    def put(self, request, *args, **kwargs):
        try:
            model = IsMember.objects.get(inv_group=kwargs['id_gi'], professor=kwargs['id_p'], member_status=True)
        except IsMember.DoesNotExist:
            return Response(f"No existe un registro en la base de datos para los datos ingresados", status=status.HTTP_404_NOT_FOUND)

        serializer = IsMemberSerializer(model, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_202_ACCEPTED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class ConsultMemberIGAPI(APIView):
    """
    Clase usada para la implementacion de, la API para consultar todos los miembros de un grupo de investigacion
    espesifico de la Universidad, esto se logra enviando el ID del Grupo de Investigacion (en ese orden) mediante
    el metodo GET
    - - - - -
    Parameters
    - - - - -
    id : int
        Referencia a un grupo de investigacion
    - - - - -
    Returned
    - - - - -
        Si el id es correcto y se encuentran resultados:
            {"Members": JsonResultado, HTTP_202_ACCEPTED}}
        Si no se encuentran resultados:
            Se mostrata un mensaje de error, HTTP_404_NOT_FOUND
    """
    def get(self, request, *args, **kwargs):
        queryset = IsMember.objects.filter(inv_group=kwargs['id'], member_status=True)
        returned = IsMemberSerializer(queryset, many=True).data
        if returned:
            return Response({"Members": returned}, status=status.HTTP_202_ACCEPTED)
        else:
            return Response(f"No existen registros en la base de datos para el ID ingresado", status=status.HTTP_404_NOT_FOUND)

class ConsultMemberProfessorAPI(APIView):
    """
    Clase usada para la implementacion de, la API para consultar si un profesor ES MIEMBRO o no de un 
    grupo de investigacion de la Universidad, esto se logra enviando el ID del Profesor mediante el metodo GET
    - - - - -
    Parameters
    - - - - -
    id : int
        Referencia a un Profesor
    - - - - -
    Returned
    - - - - -
        Si el id es correcto y se encuentran resultados:
            {"Members": JsonResultado, HTTP_202_ACCEPTED}}
        Si no se encuentran resultados:
            Se mostrata un mensaje de error, HTTP_404_NOT_FOUND
    """
    def get(self, request, *args, **kwargs):
        queryset = IsMember.objects.filter(professor=kwargs['id'], member_status=True)
        returned = IsMemberSerializer(queryset, many=True).data
        if returned:
            return Response({"Members": returned}, status=status.HTTP_202_ACCEPTED)
        else:
            queryset = IsMember.objects.filter(professor=kwargs['id'])
            returned = IsMemberSerializer(queryset, many=True).data
            if returned:
                return Response(f"El profesor no se encuentra activo", status=status.HTTP_404_NOT_FOUND)
            else:
                return Response(f"No existen registros en la base de datos para el ID ingresado", status=status.HTTP_404_NOT_FOUND)

# Trabaja
class ConsultWorksInvestGroupAPI(APIView):
    """
    Clase usada para la implementacion de, la API para consultar si un Grupo de Investigacion TRABAJA o no en un
    Area del Conocimiento espesifica de la Universidad, esto se logra enviando el ID del Grupo de Investigacion y 
    el ID del Area del Conocimiento (en ese orden) mediante el metodo GET y/o enviando la informacion 
    que se va a editar del registro del modelo "Trabaja" mediante el metodo PUT
    - - - - -
    Parameters
    - - - - -
    id_gi : int
        Referencia a un Grupo de Investigacion
    id_ac : int
        Referencia a un Area del Conocimiento
    """
    def get(self, request, *args, **kwargs):
        queryset = WorksInvestGroup.objects.filter(inv_group=kwargs['id_gi'], know_area=kwargs['id_ac'], study_status=True)
        returned = WorksInvestGroupSerializer(queryset, many=True).data
        if returned:
            return Response({"Work":returned}, status=status.HTTP_202_ACCEPTED)
        else:
            return Response(f"No existe un registro en la base de datos para los datos ingresados", status=status.HTTP_404_NOT_FOUND)
    
    def put(self, request, *args, **kwargs):
        try:
            model = WorksInvestGroup.objects.get(inv_group=kwargs['id_gi'], know_area=kwargs['id_ac'], study_status=True)
        except IsMember.DoesNotExist:
            return Response(f"No existe un registro en la base de datos para los datos ingresados", status=status.HTTP_404_NOT_FOUND)

        serializer = WorksInvestGroupSerializer(model, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_202_ACCEPTED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class ConsultWorksInvestGroup_GIAPI(APIView):
    def get(self, request, *args, **kwargs):
        queryset = WorksInvestGroup.objects.filter(inv_group=kwargs['id'], study_status=True)
        returned = WorksInvestGroupSerializer(queryset, many=True).data
        if returned:
            return Response({"Work":returned}, status=status.HTTP_202_ACCEPTED)
        else:
            return Response(f"No existe un registro en la base de datos para los datos ingresados", status=status.HTTP_404_NOT_FOUND)

# Dirige
class ConsultManageInvestGroupAPI(APIView):
    """
    Clase usada para la implementacion de, la API para consultar si un Profesor DIRIGE o no un
    Grupo de Investigacion espesifico de la Universidad, esto se logra enviando el ID del usuario del Profesor y 
    el ID del Grupo de Investigacion (en ese orden) mediante el metodo GET y/o enviando la informacion 
    que se va a editar del registro del modelo "Dirige" mediante el metodo PUT
    - - - - -
    Parameters
    - - - - -
    id_p : int
        Referencia a un Profesor
    id_gi : int
        Referencia a un Grupo de Investigacion
    """
    def get(self, request, *args, **kwargs):
        queryset = ManageInvestGroup.objects.filter(inv_group=kwargs['id_gi'], professor=kwargs['id_p'], direction_state=True)
        returned = ManageInvestGroupSerializer(queryset, many=True).data
        if returned:
            return Response({"Manage":returned}, status=status.HTTP_202_ACCEPTED)
        else:
            return Response(f"No existe un registro en la base de datos para los datos ingresados", status=status.HTTP_404_NOT_FOUND)
    
    def put(self, request, *args, **kwargs):
        try:
            model = ManageInvestGroup.objects.get(inv_group=kwargs['id_gi'], professor=kwargs['id_p'], direction_state=True)
        except IsMember.DoesNotExist:
            return Response(f"No existe un registro en la base de datos para los datos ingresados", status=status.HTTP_404_NOT_FOUND)

        serializer = ManageInvestGroupSerializer(model, data=request.data)
        if serializer.is_valid():
            if 'professor' in request.data.keys():
                oldProfessor = Professor.objects.get(id=kwargs['id_p'])
                oldProfessor.is_director_gi = False
                oldProfessor.save()
                newProfessor = Professor.objects.get(id=request.data['professor'])
                newProfessor.is_director_gi = True
                newProfessor.save()
            if 'status' in request.data.keys():
                if request.data['status'] == False:
                    oldProfessor = Professor.objects.get(id=kwargs['id_p'])
                    oldProfessor.is_director_gi = False
                    oldProfessor.save()
            serializer.save()
            return Response(serializer.data, status=status.HTTP_202_ACCEPTED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class ConsultManageInvestGroup_DirecAPI(APIView):
    """
    Clase usada para la implementacion de, la API para consultar que un profesor DIRIGE o no a un
    grupo de investigacion de la Universidad, esto se logra enviando el ID del Profesor mediante el metodo GET
    - - - - -
    Parameters
    - - - - -
    id : int
        Referencia a un Profesor
    - - - - -
    Returned
    - - - - -
        Si el id es correcto y se encuentran resultados:
            {"Manage": JsonResultado, HTTP_202_ACCEPTED}}
        Si no se encuentran resultados:
            Se mostrata un mensaje de error, HTTP_404_NOT_FOUND
    """
    def get(self, request, *args, **kwargs):
        queryset = ManageInvestGroup.objects.filter(professor=kwargs['id'], direction_state=True)
        returned = ManageInvestGroupSerializer(queryset, many=True).data
        if returned:
            return Response({"Manage": returned}, status=status.HTTP_202_ACCEPTED)
        else:
            return Response(f"No existen registros en la base de datos para el ID ingresado", status=status.HTTP_404_NOT_FOUND)

class ConsultManageInvestGroup_GIAPI(APIView):
    """
    Clase usada para la implementacion de, la API para consultar cual es el profesor que DIRIGE a un grupo de
    investigacion de la Universidad, esto se logra enviando el ID del Grupo de Investigacion mediante el metodo GET
    - - - - -
    Parameters
    - - - - -
    id : int
        Referencia a un Grupo de investigacion
    - - - - -
    Returned
    - - - - -
        Si el id es correcto y se encuentran resultados:
            {"Manage": JsonResultado, HTTP_202_ACCEPTED}}
        Si no se encuentran resultados:
            Se mostrata un mensaje de error, HTTP_404_NOT_FOUND
    """
    def get(self, request, *args, **kwargs):
        queryset = ManageInvestGroup.objects.filter(inv_group=kwargs['id'], direction_state=True)
        returned = ManageInvestGroupSerializer(queryset, many=True).data
        if returned:
            return Response({"Manage": returned}, status=status.HTTP_202_ACCEPTED)
        else:
            return Response(f"No existen registros en la base de datos para el ID ingresado", status=status.HTTP_404_NOT_FOUND)

#maneja
class ConsultManageInvestLineAPI(APIView):
    """
    Clase usada para la implementacion de, la API para consultar si un Profesor MANEJA o no una
    Linea de investigacion espesifica, esto se logra enviando el ID del Profesor y el ID de la 
    Linea de investigacion (en ese orden) mediante el metodo GET y/o enviando la informacion
    que se va a editar del registro del modelo "maneja" mediante el metodo PUT
    - - - - -
    Parameters
    - - - - -
    id_p : int
        Referencia a un Profesor
    id_i : int
        Referencia a una Linea de Investigacion de la Universidad
    """
    def get(self, request, *args, **kwargs):
        queryset = ManageInvestLine.objects.filter(inv_line=kwargs['id_i'], professor=kwargs['id_p'], analysis_state=True)
        returned = ManageInvestLineSerializer(queryset, many=True).data
        if returned:
            return Response(returned, status=status.HTTP_202_ACCEPTED)
        else:
            return Response(f"No existe un registro en la base de datos para los datos ingresados", status=status.HTTP_404_NOT_FOUND)
    
    def put(self, request, *args, **kwargs):
        try:
            model = ManageInvestLine.objects.filter(inv_line=kwargs['id_i'], professor=kwargs['id_p'], analysis_state=True)
        except ManageInvestLine.DoesNotExist:
            return Response(f"No existe un registro en la base de datos para los datos ingresados", status=status.HTTP_404_NOT_FOUND)
        
        serializer = ManageInvestLineSerializer(model, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_202_ACCEPTED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class ConsultManageInvestLine_invLineAPI(APIView):
    """
    Clase usada para la implementacion de, la API para consultar cuales Profesores MANEJAN una
    Linea de Investigacion, esto se logra enviando el ID de la Linea de Investigacion mediante el metodo GET 
    - - - - -
    Parameters
    - - - - -
    id : int
        Referencia a una Linea de Investigacion
    """
    def get(self, request, *args, **kwargs):
        queryset = ManageInvestLine.objects.filter(inv_line=kwargs['id'], analysis_state=True)
        returned = ManageInvestLineSerializer(queryset, many=True).data
        if returned:
            return Response(returned, status=status.HTTP_202_ACCEPTED)
        else:
            return Response(f"No existe un registro en la base de datos para los datos ingresados", status=status.HTTP_404_NOT_FOUND)

class ConsultManageInvestLine_profAPI(APIView):
    """
    Clase usada para la implementacion de, la API para consultar si un Profesor MANEJA una o mas
    Lineas de Investigacion, esto se logra enviando el ID del Profesor mediante el metodo GET 
    - - - - -
    Parameters
    - - - - -
    id : int
        Referencia a un Profesor
    """
    def get(self, request, *args, **kwargs):
        queryset = ManageInvestLine.objects.filter(professor=kwargs['id'], analysis_state=True)
        returned = ManageInvestLineSerializer(queryset, many=True).data
        if returned:
            return Response(returned, status=status.HTTP_202_ACCEPTED)
        else:
            return Response(f"No existe un registro en la base de datos para los datos ingresados", status=status.HTTP_404_NOT_FOUND)

# labora
class ConsultWorksDepartmAPI(APIView):
    """
    Clase usada para la implementacion de, la API para consultar si un Profesor LABORA o no un
    Departamento espesifico de la Universidad, esto se logra enviando el ID del Profesor y 
    el ID del Departamento (en ese orden) mediante el metodo GET y/o enviando la informacion 
    que se va a editar del registro del modelo "Labora" mediante el metodo PUT
    - - - - -
    Parameters
    - - - - -
    id_p : int
        Referencia a un Profesor
    id_d : int
        Referencia a un Departamento de la Universidad
    """
    def get(self, request, *args, **kwargs):
        queryset = WorksDepartm.objects.filter(professor=kwargs['id_p'], department=kwargs['id_d'])
        returned = WorksDepartmSerializer(queryset, many=True).data
        if returned:
            return Response(returned, status=status.HTTP_202_ACCEPTED)
        else:
            return Response(f"No existe un registro en la base de datos para los datos ingresados", status=status.HTTP_404_NOT_FOUND)
    
    def put(self, request, *args, **kwargs):
        try:
            model = WorksDepartm.objects.get(professor=kwargs['id_p'], department=kwargs['id_d'])
        except WorksDepartm.DoesNotExist:
            return Response(f"No existe un registro en la base de datos para los datos ingresados", status=status.HTTP_404_NOT_FOUND)
        
        serializer = WorksDepartmSerializer(model, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_202_ACCEPTED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class ConsultWorksDepartm_profAPI(APIView):
    """
    Clase usada para la implementacion de, la API para consultar si un Profesor LABORA o no un
    Departamento de la Universidad, esto se logra enviando el ID del Profesor mediante el metodo GET 
    - - - - -
    Parameters
    - - - - -
    id : int
        Referencia a un Profesor
    """
    def get(self, request, *args, **kwargs):
        queryset = WorksDepartm.objects.filter(professor=kwargs['id'], laboral_state=True)
        returned = WorksDepartmSerializer(queryset, many=True).data
        if returned:
            return Response(returned, status=status.HTTP_202_ACCEPTED)
        else:
            return Response(f"No existe un registro en la base de datos para los datos ingresados", status=status.HTTP_404_NOT_FOUND)

class ConsultWorksDepartm_depAPI(APIView):
    """
    Clase usada para la implementacion de, la API para consultar cuales Profesores LABORAN en un
    Departamento espesifico de la Universidad, esto se logra enviando el ID del Departamento mediante
    el metodo GET 
    - - - - -
    Parameters
    - - - - -
    id : int
        Referencia a un Departamento de la Universidad
    """
    def get(self, request, *args, **kwargs):
        queryset = WorksDepartm.objects.filter(department=kwargs['id'], laboral_state=True)
        returned = WorksDepartmSerializer(queryset, many=True).data
        if returned:
            return Response(returned, status=status.HTTP_202_ACCEPTED)
        else:
            return Response(f"No existe un registro en la base de datos para los datos ingresados", status=status.HTTP_404_NOT_FOUND)

# Formacion academica
class ConsultAcademicTrainingAPI(APIView):
    """
    Clase usada para la implementacion de, la API para consultar la FORMACION ACADEMICA de un
    Profesor espesifico de la Universidad, esto se logra enviando el ID de la Formacion Academica
    correspondiente mediante el metodo GET y/o enviando la informacion que se va a editar del
    registro del modelo "Formacion academica" mediante el metodo PUT
    - - - - -
    Parameters
    - - - - -
    id : int
        Referencia a una Formacion academica (Primary key del registro de la tabla)
    """
    def get(self, request, *args, **kwargs):
        queryset = AcademicTraining.objects.filter(id=kwargs['id'], professor__status=True)
        returned = AcademicTrainingSerializer(queryset, many=True).data
        if returned:
            return Response(returned, status=status.HTTP_202_ACCEPTED)
        else:
            return Response(f"No existe un registro en la base de datos para los datos ingresados", status=status.HTTP_404_NOT_FOUND)

    def put(self, request, *args, **kwargs):
        try:
            model = AcademicTraining.objects.get(id=kwargs['id'], professor__status=True)
        except AcademicTraining.DoesNotExist:
            return Response(f"No existe un registro en la base de datos para los datos ingresados", status=status.HTTP_404_NOT_FOUND)
        
        serializer = AcademicTrainingSerializer(model, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_202_ACCEPTED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class ConsultAcademicTraining_profAPI(APIView):
    """
    Clase usada para la implementacion de, la API para consultar las FORMACIONES ACADEMICAS de un
    Profesor de la Universidad, esto se logra enviando el ID del Profesor mediante el metodo GET 
    - - - - -
    Parameters
    - - - - -
    id : int
        Referencia a un Profesor
    """
    def get(self, request, *args, **kwargs):
        queryset = AcademicTraining.objects.filter(professor=kwargs['id'], professor__status=True)
        returned = AcademicTrainingSerializer(queryset, many=True).data
        if returned:
            return Response(returned, status=status.HTTP_202_ACCEPTED)
        else:
            return Response(f"No existe un registro en la base de datos para los datos ingresados", status=status.HTTP_404_NOT_FOUND)

# Coordinador
class ConsultCoordinatorAPI(APIView):
    # consultar por cuales campos para la edicion
    def get(self, request, *args, **kwargs):
        queryset = CoordinatorProgram.objects.filter(program=kwargs['prog'], academic_period=kwargs['period'])
        returned = CoordinatorProgramSerializer(queryset, many=True).data
        if returned:
            return Response(returned, status=status.HTTP_202_ACCEPTED)
        else:
            return Response(f"No existe un registro en la base de datos para los datos ingresados", status=status.HTTP_404_NOT_FOUND)

    def put(self, request, *args, **kwargs):
        try:
            model = CoordinatorProgram.objects.get(program=kwargs['prog'], academic_period=kwargs['period'])
        except CoordinatorProgram.DoesNotExist:
            return Response(f"No existe un registro en la base de datos para los datos ingresados", status=status.HTTP_404_NOT_FOUND)
        
        serializer = CoordinatorProgramSerializer(model, data=request.data)
        if serializer.is_valid():
            model.date_update = datetime.datetime.now()
            serializer.save()
            return Response(serializer.data, status=status.HTTP_202_ACCEPTED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


#endregion
