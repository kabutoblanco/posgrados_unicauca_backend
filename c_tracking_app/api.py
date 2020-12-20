from io import BytesIO

from django.db.models import Count, Q
from django.db.models.query import QuerySet
from django.db.models.signals import post_save
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView

from rest_framework import generics, status, views, viewsets
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
from reportlab.platypus.paragraph import Paragraph

from a_students_app.models import StudentProfessor

from .backends import IsDirector, IsCoordinador
from .models import *
from .serializers import *
from .signals import *

post_save.connect(save_or_send_testdirector, sender=TestDirector)
post_save.connect(save_or_send_testcoordinator, sender=TestCoordinator)
post_save.connect(update_tracking, sender=Tracking)

# - - - - - PRIMER SPRINT - - - - -
class StudentAPI(generics.RetrieveAPIView):
    """
    Retorna un estudiante con una matricula
    - - - - -
    Parameters
    - - - - -
    id_student : int
        Id correspondiente al usuario del estudiante
    - - - - -
    Permissions
    - - - - -
    IsAuthenticated
        El usuario debe enviar un token para identificarlo
    """

    # permission_classes = (IsAuthenticated, )
    serializer_class = EnrrollmentSerializer

    def get(self, request, *args, **kwargs):
        try:
            queryset = Enrrollment.objects.filter(
                student__user=kwargs['id_student'], is_active=True).order_by('-period')[0]
            return Response({"student": self.get_serializer(queryset).data})
        except:
            return Response({"student": []})


class StudentListAPI(generics.ListAPIView):
    """
    Retorna una lista de estudiantes con la matricula respectiva
    - - - - -
    Permissions
    - - - - -
    IsAuthenticated
        El usuario debe enviar un token para identificarlo
    """

    # permission_classes = (IsAuthenticated, )
    serializer_class = EnrrollmentSerializer

    def get(self, request, *args, **kwargs):
        queryset = Student.objects.filter(is_active=True)
        queryset_list = Enrrollment.objects.none()
        for student in queryset:
            queryset_list = queryset_list | Enrrollment.objects.filter(
                student=student).order_by('-period')[:1]
        return Response({"students": self.get_serializer(queryset_list, many=True).data})


class TrackingAPI(viewsets.ModelViewSet):
    """
    CRUD de seguimiento
    - - - - -
    Permissions
    - - - - -
    IsAuthenticated
        El usuario debe enviar un token para identificarlo
    IsDirector
        El usuario debe tener rol director y/o coodirector
    IsCoordinator
        El usuario debe tener rol de coordinador
    """

    # permission_classes = [IsAuthenticated, (IsDirector | IsCoordinador)]
    serializer_class = TrackingSerializer
    queryset = Tracking.objects.all()


class ActivityListStudentAPI(generics.RetrieveAPIView):
    """
    Retorna las actividades de un estudiante
    - - - - -
    Parameters
    - - - - -
    id_student : int
        Id correspondiente al usuario del estudiante
    - - - - -
    Permissions
    - - - - -
    IsAuthenticated
        El usuario debe enviar un token para identificarlo
    IsDirector
        El usuario debe tener rol director y/o coodirector
    IsCoordinator
        El usuario debe tener rol de coordinador
    """

    # permission_classes = [IsAuthenticated, (IsDirector | IsCoordinador)]
    serializer_class = ActivitySerializer

    def get(self, request, *args, **kwargs):
        queryset = Activity.objects.filter(
            student__user=kwargs['id_student'], is_active=True)
        return Response({"activities": self.get_serializer(queryset, many=True).data})
# - - - - - PRIMER SPRINT - - - - -


# - - - - - SEGUNDO SPRINT - - - - -
class DirectorStudentsAPI(generics.RetrieveAPIView):
    """
    Retorna los estudiantes de un director
    - - - - -
    Parameters
    - - - - -
    id_director : int
        Id correspondiente al usuario del director
    - - - - -
    Permissions
    - - - - -
    IsAuthenticated
        El usuario debe enviar un token para identificarlo
    IsDirector
        El usuario debe tener rol director y/o coodirector
    """

    # permission_classes = (IsAuthenticated, IsDirector)
    serializer_class = EnrrollmentSerializer

    def get(self, request, *args, **kwargs):
        queryset = StudentProfessor.objects.filter(
            is_active=True, professor__user=kwargs['id_professor']).values('student')
        list = [e['student'] for e in queryset]
        queryset = Student.objects.filter(id__in=list, is_active=True)
        queryset_list = Enrrollment.objects.none()
        for student in queryset:
            queryset_list = queryset_list | Enrrollment.objects.filter(
                student=student.id).order_by('-period')[:1]
        return Response({"students": self.get_serializer(queryset_list, many=True).data})


class DirectorActiviesAPI(generics.RetrieveAPIView):
    """
    Retorna las actividades del un director
    - - - - -
    Parameters
    - - - - -
    id_director : int
        Id correspondiente al usuario del director
    - - - - -
    Permissions
    - - - - -
    IsAuthenticated
        El usuario debe enviar un token para identificarlo
    IsDirector
        El usuario debe tener a cargo al menos un estudiante como director o coodirector
    """

    # permission_classes = (IsAuthenticated, IsDirector)
    serializer_class = ActivitySerializer

    def get(self, request, *args, **kwargs):
        queryset = ActivityProfessor.objects.filter(Q(rol=1) | Q(rol=2))
        queryset = queryset.filter(
            professor__user=kwargs['id_professor'], is_active=True).values('activity')
        list = [e['activity'] for e in queryset]
        queryset_list = Activity.objects.filter(
            is_active=True, id__in=list)
        return Response({"activities": self.get_serializer(queryset_list, many=True).data})


class CoordinatorActivitiesAPI(generics.RetrieveAPIView):
    """
    Retorna las actividades del un coordinador
    - - - - -
    Parameters
    - - - - -
    id_director : int
        Id correspondiente al usuario del coordinador
    - - - - -
    Permissions
    - - - - -
    IsAuthenticated
        El usuario debe enviar un token para identificarlo
    IsCoordinator
        El usuario debe tener a cargo al menos un estudiante como coordinador
    """

    # permission_classes = (IsAuthenticated, IsCoordinador)
    serializer_class = ActivityEnabledSerializer

    def get(self, request, *args, **kwargs):
        user = kwargs['id_professor']
        queryset = ActivityProfessor.objects.filter(
            professor__user=user, rol=3, is_active=True).values('activity')
        program = CoordinatorProgram.objects.get(professor__user=user).program
        queryset_aux = ActivityProfessor.objects.filter(activity__student__program=program, is_active=True).values('activity').annotate(Count('activity')).values('activity')
        queryset_list = []
        for activity_professor_i in queryset_aux:
            activity = Activity.objects.get(pk=activity_professor_i["activity"])      
            print(queryset.values('activity'))
            activity_serializer = self.get_serializer(activity, context=self.get_serializer_context()).data
            if activity_professor_i in queryset:        
                activity_serializer['is_enabled'] = True
            else:            
                activity_serializer['is_enabled'] = False
            queryset_list.append(activity_serializer)
        return Response({"activities": queryset_list})
# - - - - - SEGUNDO SPRINT - - - - -


# - - - - - TERCER SPRINT - - - - -
class TestDirectorAPI(viewsets.ModelViewSet):
    """
    CRUD TestDirector
    - - - - -
    Permissions
    - - - - -
    IsAuthenticated
        El usuario debe enviar un token para identificarlo
    IsDirector
        El usuario debe tener a cargo al menos un estudiante como director
    """

    # permission_classes = [IsAuthenticated, IsDirector]
    serializer_class = TestDirectorSerializer
    queryset = TestDirector.objects.all()

    def create(self, request, *args, **kwargs):
        request.data['director'] = Professor.objects.get(
            user=request.data['director']).id
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        request.data['director'] = Professor.objects.get(
            user=request.data['director']).id
        return super().update(request, *args, **kwargs)


class TestCoordinatorAPI(viewsets.ModelViewSet):
    """
    CRUD TestCoordinator
    - - - - -
    Permissions
    - - - - -
    IsAuthenticated
        El usuario debe enviar un token para identificarlo
    IsCoordinator
        El usuario debe tener a cargo al menos un estudiante como coordinador
    """

    # permission_classes = [IsAuthenticated, IsCoordinador]
    serializer_class = TestCoordinatorSerializer
    queryset = TestCoordinator.objects.all()

    def create(self, request, *args, **kwargs):
        request.data['coordinator'] = Professor.objects.get(
            user=request.data['coordinator']).id
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        request.data['coordinator'] = Professor.objects.get(
            user=request.data['coordinator']).id
        return super().update(request, *args, **kwargs)


class TestDirectorListAPI(generics.RetrieveAPIView):
    """
    Retorna todas las evaluaciones de un director
    - - - - -
    Permissions
    - - - - -
    IsAuthenticated
        El usuario debe enviar un token para identificarlo
    IsDirector
        El usuario debe tener a cargo al menos un estudiante como director
    """

    # permission_classes = [IsAuthenticated, IsDirector]
    serializer_class = TestDirectorSerializer

    def get(self, request, *args, **kwargs):
        print(kwargs['id_professor'])
        queryset = TestDirector.objects.filter(
            director__user=kwargs['id_professor'], is_active=True)
        return Response({"test_activities": self.get_serializer(queryset, many=True).data})


class TestCoordinatorListAPI(generics.RetrieveAPIView):
    """
    Retorna todas las evaluaciones de un coordinador
    - - - - -
    Permissions
    - - - - -
    IsAuthenticated
        El usuario debe enviar un token para identificarlo
    IsCoordinator
        El usuario debe tener a cargo al menos un estudiante como coordinador
    """

    # permission_classes = [IsAuthenticated, IsCoordinador]
    serializer_class = TestCoordinatorSerializer

    def get(self, request, *args, **kwargs):
        queryset = TestCoordinator.objects.filter(
            coordinator__user=kwargs['id_professor'], is_active=True)
        return Response({"test_activities": self.get_serializer(queryset, many=True).data})
# - - - - - TERCER SPRINT - - - - -


class ReportAPI(TemplateView):
    def get(self, request, *args, **kwargs):
        queryset = Student.objects.filter(is_active=True)
        queryset_list1 = Enrrollment.objects.none()
        for student in queryset:
            queryset_list1 = queryset_list1 | Enrrollment.objects.filter(state=1).values(
                'period').annotate(enrrollment=Count('state')).order_by('period')
        print(queryset_list1)
        queryset_list2 = Enrrollment.objects.none()
        for student in queryset:
            queryset_list2 = queryset_list2 | Enrrollment.objects.filter(state=3).values(
                'period').annotate(enrrollment=Count('state')).order_by('period')
        print(queryset_list2)
        queryset_list3 = Enrrollment.objects.all().values(
            'period').annotate(period_i=Count('period')).order_by('period')
        print(queryset_list3)
        cohorte_list = []
        for period in queryset_list3:
            cohorte = {"period": None, "enrrollment": 0, "graduate": 0}
            cohorte["period"] = period["period"]
            for enrrollment in queryset_list1:
                if enrrollment["period"] == cohorte["period"]:
                    cohorte["enrrollment"] = enrrollment["enrrollment"]
                    break
            for graduate in queryset_list2:
                if graduate["period"] == cohorte["period"]:
                    cohorte["graduate"] = graduate["enrrollment"]
                    break
            cohorte_list.append(cohorte)
        print(cohorte_list)

        if(kwargs["type"] == 1):
            wb = Workbook()
            count = 3
            ws = wb.active
            ws.title = 'Hoja 1'
            # Create title in the sheet
            ws['A1'].alignment = Alignment(
                horizontal="center", vertical="center")
            ws['A1'].border = Border(left=Side(border_style="thin"), right=Side(
                border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['A1'].fill = PatternFill(
                start_color='66FFCC', end_color='66FFCC', fill_type='solid')
            ws['A1'].font = Font(name='Calibri', size=14, bold=True)
            ws['A1'] = "INFORME POR COHORTE"
            # Change the characteristics of cells
            ws.merge_cells('A1:E1')
            ws.row_dimensions[1].height = 25
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            # Create header
            ws['A2'].alignment = Alignment(
                horizontal="center", vertical="center")
            ws['A2'].border = Border(left=Side(border_style="thin"), right=Side(
                border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['A2'].fill = PatternFill(
                start_color='66CFCC', end_color='66CFCC', fill_type='solid')
            ws['A2'].font = Font(name='Calibri', size=12, bold=True)
            ws['A2'] = 'No.'

            ws['B2'].alignment = Alignment(
                horizontal="center", vertical="center")
            ws['B2'].border = Border(left=Side(border_style="thin"), right=Side(
                border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['B2'].fill = PatternFill(
                start_color='66CFCC', end_color='66CFCC', fill_type='solid')
            ws['B2'].font = Font(name='Calibri', size=12, bold=True)
            ws['B2'] = 'Cohorte'

            ws['C2'].alignment = Alignment(
                horizontal="center", vertical="center")
            ws['C2'].border = Border(left=Side(border_style="thin"), right=Side(
                border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['C2'].fill = PatternFill(
                start_color='66CFCC', end_color='66CFCC', fill_type='solid')
            ws['C2'].font = Font(name='Calibri', size=12, bold=True)
            ws['C2'] = 'No. Matriculados'

            ws['D2'].alignment = Alignment(
                horizontal="center", vertical="center")
            ws['D2'].border = Border(left=Side(border_style="thin"), right=Side(
                border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['D2'].fill = PatternFill(
                start_color='66CFCC', end_color='66CFCC', fill_type='solid')
            ws['D2'].font = Font(name='Calibri', size=12, bold=True)
            ws['D2'] = 'No. Graduados'

            ws['E2'].alignment = Alignment(
                horizontal="center", vertical="center")
            ws['E2'].border = Border(left=Side(border_style="thin"), right=Side(
                border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['E2'].fill = PatternFill(
                start_color='66CFCC', end_color='66CFCC', fill_type='solid')
            ws['E2'].font = Font(name='Calibri', size=12, bold=True)
            ws['E2'] = '% Graduados'

            for q in cohorte_list:
                ws.cell(row=count, column=1).alignment = Alignment(
                    horizontal="center")
                ws.cell(row=count, column=1).border = Border(left=Side(border_style="thin"), right=Side(
                    border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                ws.cell(row=count, column=1).font = Font(
                    name='Calibri', size=12)
                ws.cell(row=count, column=1).value = count - 2

                ws.cell(row=count, column=2).alignment = Alignment(
                    horizontal="center")
                ws.cell(row=count, column=2).border = Border(left=Side(border_style="thin"), right=Side(
                    border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                ws.cell(row=count, column=2).font = Font(
                    name='Calibri', size=12)
                ws.cell(row=count, column=2).value = q["period"]

                ws.cell(row=count, column=3).alignment = Alignment(
                    horizontal="center")
                ws.cell(row=count, column=3).border = Border(left=Side(border_style="thin"), right=Side(
                    border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                ws.cell(row=count, column=3).font = Font(
                    name='Calibri', size=12)
                ws.cell(row=count, column=3).value = q["enrrollment"]

                ws.cell(row=count, column=4).alignment = Alignment(
                    horizontal="center")
                ws.cell(row=count, column=4).border = Border(left=Side(border_style="thin"), right=Side(
                    border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                ws.cell(row=count, column=4).font = Font(
                    name='Calibri', size=12)
                ws.cell(row=count, column=4).value = q["graduate"]

                ws.cell(row=count, column=5).alignment = Alignment(
                    horizontal="center")
                ws.cell(row=count, column=5).border = Border(left=Side(border_style="thin"), right=Side(
                    border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                ws.cell(row=count, column=5).font = Font(
                    name='Calibri', size=12)
                ws.cell(row=count, column=5).value = str(
                    (int(q["graduate"]) / (int(q["graduate"]) + int(q["enrrollment"]))) * 100) + "%"

                count += 1

            # Name file
            filename = "reporte_cohortes.xlsx"
            # The define the type of response to be give
            response = HttpResponse(content_type="application/ms-excel")
            content = "attachment; filename = {0}".format(filename)
            response["Content-Disposition"] = content
            wb.save(response)

        if(kwargs["type"] == 2):  # Format pdf request.data["tipo"]==2
            filename = "reporte_cohortes.pdf"
            # The define the type of response to be give
            response = HttpResponse(content_type="application/pdf")
            content = "attachment; filename = {0}".format(filename)
            response["Content-Disposition"] = content
            # Create the PDF object, using the Bytesobject as its "file"
            buffer = BytesIO()
            c = canvas.Canvas(buffer, pagesize=A4)

            # Header
            c.setLineWidth(.3)
            c.setFont('Helvetica', 22)
            c.drawString(30, 750, 'Reporte por cohorte')
            c.setFont('Helvetica', 12)
            c.drawString(30, 735, 'Report')
            c.setFont('Helvetica-Bold', 12)
            c.drawString(460, 750, ' cohorte')
            c.line(460, 747, 560, 747)

            # Table header
            styles = getSampleStyleSheet()
            styleBH = styles["Normal"]
            styleBH.alignment = 1  # TA_CENTER is 1
            styleBH.fontSize = 10

            codigo = Paragraph('''No.''', styleBH)
            cohorte = Paragraph('''Cohorte''', styleBH)
            no_matriculados = Paragraph('''No. Matriculados''', styleBH)
            no_graduados = Paragraph('''No. Graduados''', styleBH)
            por_matriculados = Paragraph('''% Graduados''', styleBH)
            data = []
            data.append([codigo, cohorte, no_matriculados,
                         no_graduados, por_matriculados])

            # Table
            styleN = styles["BodyText"]
            styleN.alignment = 1  # TA_CENTER is 1
            styleN.fontSize = 7

            high = 650
            count = 1
            for q in cohorte_list:
                data.append([count,
                             q["period"],
                             q["enrrollment"],
                             q["graduate"],
                             str((
                                 int(q["graduate"]) / (int(q["graduate"]) + int(q["enrrollment"]))) * 100) + "%"
                             ])
                high = high - 18
                count += 1

            # Table zise
            width, height = A4
            table = Table(data, colWidths=[2.7*cm, 2.2*cm, 5*cm, 5*cm, 4*cm])
            table.setStyle(TableStyle([('INNERGRID', (0, 0), (-1, -1), 0.25,
                                        colors.black), ('BOX', (0, 0), (-1, -1), 0.25, colors.black)]))
            # Pdf size
            table.wrapOn(c, width, height)
            table.drawOn(c, 30, high)
            c.showPage()

            c.save()
            pdf = buffer.getvalue()
            buffer.close()
            response.write(pdf)

        return response
